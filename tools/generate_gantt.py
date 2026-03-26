#!/usr/bin/env python3
"""
Comprehensive Project-Management Gantt Chart Excel Generator
============================================================
Generates a feature-rich, CONFIG-driven workbook with:
  1. CONFIG          – single source of truth (all colours, types, statuses,
                       view settings, holidays, etc.)
  2. RESOURCES       – resource register (rates reference CONFIG hours/day)
  3. TASKS           – task entry with dropdowns driven by CONFIG named ranges
  4. GANTT_CHART     – visual weekly Gantt with CF bars, milestones, today-marker
  5. RESOURCE_ALLOC  – resource × task allocation matrix with over-alloc alert
  6. TIMESHEET       – monthly hours per resource per task
  7. COST_ANALYSIS   – budget vs estimated cost roll-up
  8. CHANGE_LOG      – audit trail
  9. INSTRUCTIONS    – usage guide

Usage:
    pip install openpyxl
    python3 generate_gantt.py
"""

import datetime
import os

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# OUTPUT
# ---------------------------------------------------------------------------
OUTPUT_DIR = os.path.join(os.path.dirname(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "project_gantt.xlsx")

# ---------------------------------------------------------------------------
# DEFAULT SEED DATA  (the user overrides everything in the CONFIG sheet)
# ---------------------------------------------------------------------------
PROJECT_NAME = "Project Alpha"
PROJECT_MANAGER = "Alice Johnson"
PROJECT_START = datetime.date(2025, 1, 6)   # Monday
PROJECT_END = datetime.date(2025, 12, 31)
WORKING_HOURS_PER_DAY = 8
CURRENCY = "USD"
GANTT_VIEW = "Weekly"

# (name, bar-hex, row-bg-hex)
ITEM_TYPES = [
    ("Epic",        "4472C4", "DDEEFF"),
    ("Feature",     "ED7D31", "FDE9D9"),
    ("Task",        "70AD47", "E2EFDA"),
    ("Milestone",   "FF0000", "FFE0E0"),
    ("Deliverable", "7030A0", "EBD6FF"),
]

# (name, bg-hex)
STATUSES = [
    ("Not Started", "D9D9D9"),
    ("Planned",     "BDD7EE"),
    ("In Progress", "FFEB9C"),
    ("Completed",   "C6EFCE"),
    ("Delayed",     "FFC7CE"),
    ("On Hold",     "FCE4D6"),
    ("Cancelled",   "808080"),
]

# (name, bg-hex)
PRIORITIES = [
    ("Low",      "C6EFCE"),
    ("Medium",   "FFEB9C"),
    ("High",     "FFC7CE"),
    ("Critical", "FF0000"),
]

# (id, name, role, dept, avail%, hourly-rate)
RESOURCES = [
    ("R01", "Alice Johnson",  "Project Manager",   "PMO",         100, 110),
    ("R02", "Bob Smith",      "Sr. Developer",     "Engineering",  75, 120),
    ("R03", "Carol Williams", "UX Designer",       "Design",      100,  90),
    ("R04", "David Brown",    "QA Engineer",       "Quality",      50,  80),
    ("R05", "Eve Davis",      "Business Analyst",  "Business",    100,  95),
    ("R06", "Frank Miller",   "Backend Dev",       "Engineering", 100, 115),
    ("R07", "Grace Wilson",   "Frontend Dev",      "Engineering",  75, 100),
    ("R08", "Henry Moore",    "DevOps Engineer",   "Operations",   50, 130),
]

# (id, wbs, name, type, status, priority, start_offset_days, duration_days, progress%, [resource_ids])
TASKS = [
    (1,  "1",     "Project Kickoff",           "Epic",        "Completed",  "High",     0,   5,   100, ["R01","R05"]),
    (2,  "1.1",   "Requirements Gathering",    "Feature",     "Completed",  "High",     0,   14,  100, ["R01","R05"]),
    (3,  "1.1.1", "Stakeholder Interviews",    "Task",        "Completed",  "Medium",   0,   7,   100, ["R01"]),
    (4,  "1.1.2", "Document Requirements",     "Task",        "Completed",  "Medium",   7,   7,   100, ["R05"]),
    (5,  "1.2",   "Requirements Sign-off",     "Milestone",   "Completed",  "Critical", 14,  1,   100, ["R01"]),
    (6,  "2",     "Design Phase",              "Epic",        "Completed",  "High",     14,  28,  100, ["R03","R05"]),
    (7,  "2.1",   "UI/UX Design",              "Feature",     "Completed",  "High",     14,  21,  100, ["R03"]),
    (8,  "2.1.1", "Wireframes",                "Task",        "Completed",  "Medium",   14,  10,  100, ["R03"]),
    (9,  "2.1.2", "Prototypes",                "Task",        "Completed",  "Medium",   24,  11,  100, ["R03"]),
    (10, "2.2",   "Architecture Design",       "Feature",     "Completed",  "High",     14,  14,  100, ["R02","R06"]),
    (11, "2.3",   "Design Review",             "Milestone",   "Completed",  "Critical", 42,  1,   100, ["R01"]),
    (12, "3",     "Development Sprint 1",      "Epic",        "In Progress","High",     42,  42,   60, ["R02","R06","R07"]),
    (13, "3.1",   "Backend API Development",   "Feature",     "In Progress","High",     42,  28,   70, ["R02","R06"]),
    (14, "3.1.1", "User Authentication",       "Task",        "Completed",  "High",     42,  10,  100, ["R06"]),
    (15, "3.1.2", "Data Models",               "Task",        "Completed",  "Medium",   52,  10,  100, ["R02"]),
    (16, "3.1.3", "REST API Endpoints",        "Task",        "In Progress","High",     62,  8,    50, ["R02","R06"]),
    (17, "3.2",   "Frontend Development",      "Feature",     "In Progress","High",     56,  28,   40, ["R07"]),
    (18, "3.2.1", "Component Library",         "Task",        "Completed",  "Medium",   56,  14,  100, ["R07"]),
    (19, "3.2.2", "Dashboard Views",           "Task",        "In Progress","High",     70,  14,   30, ["R07"]),
    (20, "3.3",   "Sprint 1 Review",           "Milestone",   "Planned",    "Critical", 84,  1,     0, ["R01"]),
    (21, "4",     "Development Sprint 2",      "Epic",        "Planned",    "High",     84,  42,    0, ["R02","R06","R07"]),
    (22, "4.1",   "Integration Development",   "Feature",     "Planned",    "High",     84,  21,    0, ["R02","R07"]),
    (23, "4.2",   "Performance Optimization",  "Feature",     "Planned",    "Medium",  105,  21,    0, ["R06"]),
    (24, "4.3",   "Sprint 2 Deliverable",      "Deliverable", "Planned",    "Critical",126,  1,     0, ["R01"]),
    (25, "5",     "Testing & QA",              "Epic",        "Planned",    "High",    120,  28,    0, ["R04","R02"]),
    (26, "5.1",   "Unit Testing",              "Feature",     "Planned",    "Medium",  120,  14,    0, ["R02"]),
    (27, "5.2",   "Integration Testing",       "Feature",     "Planned",    "High",    134,  14,    0, ["R04"]),
    (28, "5.3",   "UAT",                       "Feature",     "Planned",    "High",    148,  10,    0, ["R01","R04","R05"]),
    (29, "5.4",   "Testing Sign-off",          "Milestone",   "Planned",    "Critical",158,  1,     0, ["R01"]),
    (30, "6",     "Deployment",                "Epic",        "Planned",    "Critical",158,  14,    0, ["R08","R02"]),
    (31, "6.1",   "Staging Deployment",        "Task",        "Planned",    "High",    158,  7,     0, ["R08"]),
    (32, "6.2",   "Production Deployment",     "Deliverable", "Planned",    "Critical",165,  3,     0, ["R08","R01"]),
    (33, "6.3",   "Go Live",                   "Milestone",   "Planned",    "Critical",168,  1,     0, ["R01"]),
    (34, "7",     "Post-Launch Support",        "Feature",     "Planned",    "Medium",  169,  14,    0, ["R02","R04"]),
    (35, "7.1",   "Project Closure Report",    "Deliverable", "Planned",    "Medium",  183,  7,     0, ["R01","R05"]),
]

# Holidays
HOLIDAYS = [
    (datetime.date(2025, 1,  1),  "New Year's Day"),
    (datetime.date(2025, 4, 18),  "Good Friday"),
    (datetime.date(2025, 5, 26),  "Memorial Day"),
    (datetime.date(2025, 7,  4),  "Independence Day"),
    (datetime.date(2025, 9,  1),  "Labor Day"),
    (datetime.date(2025, 11, 27), "Thanksgiving"),
    (datetime.date(2025, 12, 25), "Christmas Day"),
]

# ===========================================================================
# STYLE HELPERS
# ===========================================================================

def fill(hex6, ftype="solid"):
    return PatternFill(start_color=hex6, end_color=hex6, fill_type=ftype)

def border(l=True, r=True, t=True, b=True, style="thin"):
    s = Side(style=style)
    n = Side(style=None)
    return Border(left=s if l else n, right=s if r else n,
                  top=s if t else n,  bottom=s if b else n)

def hdr(cell, bg="1F4E79", fg="FFFFFF", sz=10, wrap=True):
    cell.fill = fill(bg)
    cell.font = Font(bold=True, color=fg, size=sz)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=wrap)
    cell.border = border()

def lbl(cell):
    """Label (right-aligned) for CONFIG key column."""
    cell.font = Font(bold=True, size=10, color="1F4E79")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = border()
    cell.fill = fill("EBF3FB")

def inp(cell):
    """Editable input cell (yellow)."""
    cell.fill = fill("FFFFC0")
    cell.font = Font(size=10)
    cell.border = border()
    cell.alignment = Alignment(horizontal="left", vertical="center")

def sec(cell, bg="2E75B6", fg="FFFFFF"):
    """Section header spanning cell."""
    cell.fill = fill(bg)
    cell.font = Font(bold=True, color=fg, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border()

def dat(cell, bg="FFFFFF"):
    """Ordinary data cell."""
    cell.fill = fill(bg)
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border()

def cw(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ===========================================================================
# NAMED-RANGE HELPER
# ===========================================================================

def nr(wb, name, sheet, cell_range):
    """Add a workbook-level named range."""
    dn = DefinedName(name, attr_text=f"'{sheet}'!{cell_range}")
    wb.defined_names[name] = dn

# ===========================================================================
# SHEET 1 — CONFIG
# ===========================================================================

def build_config(wb):
    ws = wb.create_sheet("CONFIG")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "1F4E79"

    # ── Master title ────────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    ws["A1"].value = "⚙  PROJECT GANTT – CONFIGURATION CENTRE"
    ws["A1"].fill = fill("1F4E79")
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:N2")
    ws["A2"].value = (
        "Yellow cells are editable.  All other sheets read their settings "
        "from this page via named ranges.  Do NOT rename this sheet."
    )
    ws["A2"].fill = fill("BDD7EE")
    ws["A2"].font = Font(italic=True, size=10, color="1F4E79")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── SECTION A: Project Information (cols A-C, rows 4-17) ────────────────
    ws.merge_cells("A4:C4")
    ws["A4"].value = "A  |  PROJECT INFORMATION"
    sec(ws["A4"], "2E75B6")
    ws.row_dimensions[4].height = 22

    A_DATA = [
        # (row, label, value, number_format, has_dropdown, dv_formula)
        (5,  "Project Name",          PROJECT_NAME,              "@",           False, None),
        (6,  "Project Manager",       PROJECT_MANAGER,           "@",           False, None),
        (7,  "Version",               "1.0",                     "@",           False, None),
        (8,  "Department / PMO",      "PMO",                     "@",           False, None),
        (9,  "Start Date",            PROJECT_START,             "YYYY-MM-DD",  False, None),
        (10, "End Date",              PROJECT_END,               "YYYY-MM-DD",  False, None),
        (11, "Working Hours / Day",   WORKING_HOURS_PER_DAY,     "0.0",         False, None),
        (12, "Currency",              CURRENCY,                  "@",           True,  '"USD,EUR,GBP,TRY,JPY,AUD,CAD"'),
        (13, "Gantt View Type",       GANTT_VIEW,                "@",           True,  '"Daily,Weekly,Monthly"'),
        (14, "Total Budget",          500000,                    '#,##0.00',    False, None),
        (15, "Contingency %",         10,                        "0.0",         False, None),
        (16, "Daily Rate Multiplier", 1.0,                       "0.00",        False, None),
        (17, "Show Progress Bar",     1,                         '"Yes";"Yes";"No"', True, '"1,0"'),
    ]

    for row, label, value, fmt, has_dv, dv_f in A_DATA:
        lc = ws[f"A{row}"]
        lc.value = label
        lbl(lc)

        vc = ws[f"B{row}"]
        vc.value = value
        inp(vc)
        vc.number_format = fmt

        if has_dv and dv_f:
            dv = DataValidation(type="list", formula1=dv_f, showDropDown=False)
            dv.sqref = f"B{row}"
            ws.add_data_validation(dv)

    # Working days (col C)
    ws.merge_cells("C4:C4")   # already merged above; working days in E4 area
    ws.merge_cells("E4:G4")
    ws["E4"].value = "WORKING DAYS"
    sec(ws["E4"], "2E75B6")

    WD = [("E5","Mon","F5",1), ("E6","Tue","F6",1), ("E7","Wed","F7",1),
          ("E8","Thu","F8",1), ("E9","Fri","F9",1), ("E10","Sat","F10",0),
          ("E11","Sun","F11",0)]
    dv_wd = DataValidation(type="list", formula1='"1,0"', showDropDown=False)
    dv_wd.sqref = "F5:F11"
    ws.add_data_validation(dv_wd)
    for laddr, lv, vaddr, vval in WD:
        lc = ws[laddr]; lc.value = lv; lbl(lc)
        vc = ws[vaddr]; vc.value = vval; inp(vc)
        vc.number_format = '"Yes";"Yes";"No"'
        vc.alignment = Alignment(horizontal="center")

    # ── SECTION B: Item Types (cols I-N, rows 4-11) ────────────────────────
    TYPE_ROW = 4
    ws.merge_cells("I4:N4")
    ws["I4"].value = "B  |  ITEM TYPES  ← bar colours drive the Gantt chart"
    sec(ws["I4"], "2E75B6")

    for ci, htext in zip(["I","J","K","L","M","N"],
                         ["#","Name","Bar Colour\n(HEX)","Text Colour\n(HEX)","Row BG\n(HEX)","Preview"]):
        hdr(ws[f"{ci}{TYPE_ROW+1}"], "4472C4", sz=9)
        ws[f"{ci}{TYPE_ROW+1}"].value = htext

    TYPE_NAME_ROWS = []
    for i, (tname, bar_hex, bg_hex) in enumerate(ITEM_TYPES, 1):
        row = TYPE_ROW + 1 + i
        TYPE_NAME_ROWS.append(row)
        txt_hex = "FFFFFF" if tname in ("Epic", "Milestone") else "000000"
        ws[f"I{row}"].value = i
        ws[f"J{row}"].value = tname
        ws[f"K{row}"].value = bar_hex   # editable hex → drives CF
        ws[f"L{row}"].value = txt_hex
        ws[f"M{row}"].value = bg_hex
        pv = ws[f"N{row}"]
        pv.value = tname
        pv.fill = fill(bar_hex)
        pv.font = Font(bold=True, size=9, color=txt_hex)
        pv.alignment = Alignment(horizontal="center", vertical="center")
        pv.border = border()
        for ci in ["I","J","K","L","M"]:
            c = ws[f"{ci}{row}"]
            c.border = border(); c.font = Font(size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill("F2F2F2")
        inp(ws[f"K{row}"]); inp(ws[f"L{row}"]); inp(ws[f"M{row}"])
        ws[f"J{row}"].fill = fill(bar_hex)
        ws[f"J{row}"].font = Font(bold=True, size=10, color=txt_hex)

    TYPE_FIRST_ROW = TYPE_ROW + 2
    TYPE_LAST_ROW  = TYPE_ROW + 1 + len(ITEM_TYPES)

    # ── SECTION C: Statuses (cols I-N, rows 13+) ───────────────────────────
    STA_START = TYPE_LAST_ROW + 2
    ws.merge_cells(f"I{STA_START}:N{STA_START}")
    ws[f"I{STA_START}"].value = "C  |  STATUSES  ← colours drive conditional formatting"
    sec(ws[f"I{STA_START}"], "375623")

    for ci, htext in zip(["I","J","K","L","M","N"],
                         ["#","Name","Colour\n(HEX)","Font Colour\n(HEX)","—","Preview"]):
        hdr(ws[f"{ci}{STA_START+1}"], "70AD47", sz=9)
        ws[f"{ci}{STA_START+1}"].value = htext

    STA_FIRST = STA_START + 2
    for i, (sname, s_hex) in enumerate(STATUSES, 1):
        row = STA_START + 1 + i
        ws[f"I{row}"].value = i
        ws[f"J{row}"].value = sname
        ws[f"K{row}"].value = s_hex
        ws[f"L{row}"].value = "000000"
        ws[f"M{row}"].value = ""
        pv = ws[f"N{row}"]
        pv.value = sname; pv.fill = fill(s_hex)
        pv.font = Font(size=9); pv.border = border()
        pv.alignment = Alignment(horizontal="center", vertical="center")
        for ci in ["I","J","K","L","M"]:
            c = ws[f"{ci}{row}"]
            c.border = border(); c.font = Font(size=10)
            c.alignment = Alignment(horizontal="center"); c.fill = fill("F2F2F2")
        inp(ws[f"K{row}"]); inp(ws[f"L{row}"])
        ws[f"J{row}"].fill = fill(s_hex); ws[f"J{row}"].font = Font(size=10)

    STA_LAST = STA_START + 1 + len(STATUSES)

    # ── SECTION D: Priorities ───────────────────────────────────────────────
    PRI_START = STA_LAST + 2
    ws.merge_cells(f"I{PRI_START}:N{PRI_START}")
    ws[f"I{PRI_START}"].value = "D  |  PRIORITIES"
    sec(ws[f"I{PRI_START}"], "7030A0")

    for ci, htext in zip(["I","J","K","L","M","N"],
                         ["#","Name","Colour\n(HEX)","Font Colour\n(HEX)","—","Preview"]):
        hdr(ws[f"{ci}{PRI_START+1}"], "9DC3E6", sz=9)
        ws[f"{ci}{PRI_START+1}"].value = htext

    PRI_FIRST = PRI_START + 2
    for i, (pname, p_hex) in enumerate(PRIORITIES, 1):
        row = PRI_START + 1 + i
        ws[f"I{row}"].value = i
        ws[f"J{row}"].value = pname
        ws[f"K{row}"].value = p_hex
        ws[f"L{row}"].value = "000000"
        ws[f"M{row}"].value = ""
        pv = ws[f"N{row}"]
        pv.value = pname; pv.fill = fill(p_hex)
        pv.font = Font(size=9); pv.border = border()
        pv.alignment = Alignment(horizontal="center", vertical="center")
        for ci in ["I","J","K","L","M"]:
            c = ws[f"{ci}{row}"]
            c.border = border(); c.font = Font(size=10)
            c.alignment = Alignment(horizontal="center"); c.fill = fill("F2F2F2")
        inp(ws[f"K{row}"]); inp(ws[f"L{row}"])
        ws[f"J{row}"].fill = fill(p_hex); ws[f"J{row}"].font = Font(size=10)

    PRI_LAST = PRI_START + 1 + len(PRIORITIES)

    # ── SECTION E: Gantt Display Settings (cols A-C, rows 19+) ─────────────
    GDS = 19
    ws.merge_cells(f"A{GDS}:C{GDS}")
    ws[f"A{GDS}"].value = "E  |  GANTT DISPLAY SETTINGS"
    sec(ws[f"A{GDS}"], "C55A11")

    DISP = [
        (GDS+1,  "Show Today Marker",    1,          '"Yes";"Yes";"No"'),
        (GDS+2,  "Today Marker Colour",  "FF0000",   None),
        (GDS+3,  "Weekend Colour (HEX)", "F2F2F2",   None),
        (GDS+4,  "Milestone Symbol",     "◆",        None),
        (GDS+5,  "Deliverable Symbol",   "★",        None),
        (GDS+6,  "Gantt Row Height",     18,         None),
        (GDS+7,  "Show Week Numbers",    1,          '"Yes";"Yes";"No"'),
        (GDS+8,  "Bar Opacity Label",    "100%",     None),
        (GDS+9,  "Grid Line Colour",     "D9D9D9",   None),
        (GDS+10, "Highlight Critical",   1,          '"Yes";"Yes";"No"'),
    ]
    for row, lab, val, dv_f in DISP:
        lc = ws[f"A{row}"]; lc.value = lab; lbl(lc)
        vc = ws[f"B{row}"]; vc.value = val; inp(vc)
        if dv_f:
            dv = DataValidation(type="list", formula1=dv_f, showDropDown=False)
            dv.sqref = f"B{row}"
            ws.add_data_validation(dv)

    # ── SECTION F: Holidays (cols E-G, rows 19+) ───────────────────────────
    ws.merge_cells(f"E{GDS}:G{GDS}")
    ws[f"E{GDS}"].value = "F  |  PUBLIC HOLIDAYS"
    sec(ws[f"E{GDS}"], "C55A11")

    hdr(ws[f"E{GDS+1}"], "ED7D31"); ws[f"E{GDS+1}"].value = "Date"
    hdr(ws[f"F{GDS+1}"], "ED7D31"); ws[f"F{GDS+1}"].value = "Description"
    hdr(ws[f"G{GDS+1}"], "ED7D31"); ws[f"G{GDS+1}"].value = "Type"

    HOL_FIRST = GDS + 2
    for j, (hdate, hdesc) in enumerate(HOLIDAYS):
        r = HOL_FIRST + j
        for ci, val, fmt in [("E", hdate, "YYYY-MM-DD"), ("F", hdesc, "@"),
                              ("G", "Public Holiday", "@")]:
            c = ws[f"{ci}{r}"]
            c.value = val; c.number_format = fmt
            c.border = border(); c.font = Font(size=10)
            c.fill = fill("FFF2CC")
            c.alignment = Alignment(vertical="center")

    HOL_LAST = HOL_FIRST + len(HOLIDAYS) - 1

    # ── Column widths ────────────────────────────────────────────────────────
    for idx, w in [(1,22),(2,22),(3,2),(4,2),(5,18),(6,30),(7,16),
                   (8,2),(9,5),(10,18),(11,16),(12,14),(13,14),(14,14)]:
        cw(ws, idx, w)

    # ── Named Ranges ────────────────────────────────────────────────────────
    nr(wb, "CFG_ProjectName",    "CONFIG", "$B$5")
    nr(wb, "CFG_StartDate",      "CONFIG", "$B$9")
    nr(wb, "CFG_EndDate",        "CONFIG", "$B$10")
    nr(wb, "CFG_HoursPerDay",    "CONFIG", "$B$11")
    nr(wb, "CFG_Currency",       "CONFIG", "$B$12")
    nr(wb, "CFG_GanttView",      "CONFIG", "$B$13")
    nr(wb, "CFG_Budget",         "CONFIG", "$B$14")
    nr(wb, "CFG_Contingency",    "CONFIG", "$B$15")
    nr(wb, "CFG_TodayColour",    "CONFIG", "$B$20")
    nr(wb, "CFG_WeekendColour",  "CONFIG", "$B$21")
    nr(wb, "CFG_MilestoneSym",   "CONFIG", "$B$22")
    nr(wb, "CFG_DeliverSym",     "CONFIG", "$B$23")
    nr(wb, "CFG_RowHeight",      "CONFIG", "$B$24")

    # Lists used for data-validation in other sheets
    nr(wb, "ItemTypeNames", "CONFIG",
       f"$J${TYPE_FIRST_ROW}:$J${TYPE_LAST_ROW}")
    nr(wb, "StatusNames",   "CONFIG",
       f"$J${STA_FIRST}:$J${STA_LAST}")
    nr(wb, "PriorityNames", "CONFIG",
       f"$J${PRI_FIRST}:$J${PRI_LAST}")
    nr(wb, "HolidayDates",  "CONFIG",
       f"$E${HOL_FIRST}:$E${HOL_LAST}")

    # Store row indices so other builders can reference the colour cells
    config_meta = {
        "type_first_row":  TYPE_FIRST_ROW,
        "type_last_row":   TYPE_LAST_ROW,
        "sta_first_row":   STA_FIRST,
        "sta_last_row":    STA_LAST,
        "pri_first_row":   PRI_FIRST,
        "pri_last_row":    PRI_LAST,
        "hol_first_row":   HOL_FIRST,
        "hol_last_row":    HOL_LAST,
    }
    return ws, config_meta


# ===========================================================================
# SHEET 2 — RESOURCES
# ===========================================================================

def build_resources(wb):
    ws = wb.create_sheet("RESOURCES")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "2E75B6"
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:M1")
    ws["A1"].value = "👥  RESOURCE REGISTER"
    ws["A1"].fill = fill("2E75B6")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"].value = (
        "Define all project resources.  "
        "Daily Rate = Hourly Rate × CONFIG!B11 (Working Hours/Day).  "
        "Resources populate dropdowns in TASKS and RESOURCE_ALLOC."
    )
    ws["A2"].fill = fill("DEEAF1")
    ws["A2"].font = Font(italic=True, size=9, color="2E75B6")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    COLS = [
        ("A","ID",         7), ("B","Full Name",  24), ("C","Role",       22),
        ("D","Department", 18), ("E","Email",      28), ("F","Avail. %",  10),
        ("G","Hourly Rate",12), ("H","Daily Rate", 12), ("I","Colour",     10),
        ("J","Skills",     28), ("K","Location",   14), ("L","Contract",   14),
        ("M","Notes",      28),
    ]
    for ci, ht, w in COLS:
        c = ws[f"{ci}3"]; c.value = ht; hdr(c, "2E75B6")
        ws.column_dimensions[ci].width = w
    ws.row_dimensions[3].height = 22

    RES_COLORS = ["4472C4","ED7D31","A9D18E","FF0000","FFC000",
                  "7030A0","00B0F0","92D050"]
    for i, (rid, name, role, dept, avail, rate) in enumerate(RESOURCES, 1):
        row = 3 + i
        rc = RES_COLORS[(i-1) % len(RES_COLORS)]
        bg = "F5F9FF" if i % 2 else "FFFFFF"
        email = f"{name.lower().replace(' ','.')}@company.com"
        vals = [
            ("A", rid,                               "@",       "center"),
            ("B", name,                              "@",       "left"),
            ("C", role,                              "@",       "left"),
            ("D", dept,                              "@",       "left"),
            ("E", email,                             "@",       "left"),
            ("F", avail / 100,                       "0%",      "center"),
            ("G", rate,                              '#,##0.00',"right"),
            ("H", f"=G{row}*CFG_HoursPerDay",        '#,##0.00',"right"),
            ("I", "",                                "@",       "center"),
            ("J", "",                                "@",       "left"),
            ("K", "HQ",                              "@",       "center"),
            ("L", "Full-time",                       "@",       "center"),
            ("M", "",                                "@",       "left"),
        ]
        for ci, val, fmt, align in vals:
            c = ws[f"{ci}{row}"]
            c.value = val; c.number_format = fmt; c.fill = fill(bg)
            c.border = border(); c.font = Font(size=10)
            c.alignment = Alignment(horizontal=align, vertical="center")
        ws[f"I{row}"].fill = fill(rc)  # colour swatch

    LAST_DATA = 3 + len(RESOURCES)
    # Totals
    ws[f"A{LAST_DATA+1}"].value = "AVG / TOTAL"
    ws[f"A{LAST_DATA+1}"].font = Font(bold=True, size=10)
    ws[f"A{LAST_DATA+1}"].fill = fill("BDD7EE")
    ws[f"F{LAST_DATA+1}"].value = f"=AVERAGE(F4:F{LAST_DATA})"
    ws[f"F{LAST_DATA+1}"].number_format = "0%"
    ws[f"F{LAST_DATA+1}"].font = Font(bold=True); ws[f"F{LAST_DATA+1}"].fill = fill("BDD7EE")
    ws[f"G{LAST_DATA+1}"].value = f"=AVERAGE(G4:G{LAST_DATA})"
    ws[f"G{LAST_DATA+1}"].number_format = "#,##0.00"
    ws[f"G{LAST_DATA+1}"].font = Font(bold=True); ws[f"G{LAST_DATA+1}"].fill = fill("BDD7EE")

    # Named ranges
    nr(wb, "ResourceIDs",   "RESOURCES", f"$A$4:$A${LAST_DATA}")
    nr(wb, "ResourceNames", "RESOURCES", f"$B$4:$B${LAST_DATA}")
    nr(wb, "ResourceRoles", "RESOURCES", f"$C$4:$C${LAST_DATA}")
    nr(wb, "ResourceRates", "RESOURCES", f"$H$4:$H${LAST_DATA}")   # daily rate
    nr(wb, "ResourceAvail", "RESOURCES", f"$F$4:$F${LAST_DATA}")
    return ws


# ===========================================================================
# SHEET 3 — TASKS
# ===========================================================================

def build_tasks(wb):
    ws = wb.create_sheet("TASKS")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "375623"
    ws.freeze_panes = "D4"

    ws.merge_cells("A1:T1")
    ws["A1"].value = "📋  TASK REGISTER"
    ws["A1"].fill = fill("375623")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:T2")
    ws["A2"].value = (
        "Enter tasks here.  Dropdowns are driven by CONFIG named ranges.  "
        "Duration = End − Start + 1.  Up to 3 resources per task."
    )
    ws["A2"].fill = fill("E2EFDA")
    ws["A2"].font = Font(italic=True, size=9, color="375623")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    COLS = [
        ("A","ID",          6), ("B","WBS",          9), ("C","Level",        7),
        ("D","Task Name",  36), ("E","Type",         14), ("F","Status",      14),
        ("G","Priority",   11), ("H","Start Date",   13), ("I","End Date",    13),
        ("J","Duration\n(days)",10),("K","Progress\n%",9),
        ("L","Resource 1", 18), ("M","R1 %",          8),
        ("N","Resource 2", 18), ("O","R2 %",          8),
        ("P","Resource 3", 18), ("Q","R3 %",          8),
        ("R","Budget",     13), ("S","Depends On",   12), ("T","Notes",       30),
    ]
    for ci, ht, w in COLS:
        c = ws[f"{ci}3"]; c.value = ht; hdr(c, "375623")
        ws.column_dimensions[ci].width = w
    ws.row_dimensions[3].height = 30

    MAX_ROW = 300

    # Data validations — all lists sourced from CONFIG named ranges
    dv_type   = DataValidation(type="list", formula1="ItemTypeNames",  showDropDown=False)
    dv_status = DataValidation(type="list", formula1="StatusNames",    showDropDown=False)
    dv_prio   = DataValidation(type="list", formula1="PriorityNames",  showDropDown=False)
    dv_res    = DataValidation(type="list", formula1="ResourceNames",  showDropDown=False)
    dv_pct    = DataValidation(type="decimal", operator="between",
                               formula1="0", formula2="1",
                               error="Enter 0–1 (e.g. 0.5 = 50%)",
                               errorTitle="Invalid %")
    dv_type.sqref   = f"E4:E{MAX_ROW}"
    dv_status.sqref = f"F4:F{MAX_ROW}"
    dv_prio.sqref   = f"G4:G{MAX_ROW}"
    dv_res.sqref    = f"L4:L{MAX_ROW} N4:N{MAX_ROW} P4:P{MAX_ROW}"
    dv_pct.sqref    = f"M4:M{MAX_ROW} O4:O{MAX_ROW} Q4:Q{MAX_ROW}"
    for dv in [dv_type, dv_status, dv_prio, dv_res, dv_pct]:
        ws.add_data_validation(dv)

    # Populate sample tasks
    type_bg = {t[0]: t[2] for t in ITEM_TYPES}

    for (tid, wbs, name, ttype, status, priority,
         offset, dur, prog, resources) in TASKS:
        row = 3 + tid
        start = PROJECT_START + datetime.timedelta(days=offset)
        end   = start + datetime.timedelta(days=dur - 1)
        indent = wbs.count(".")
        bg = type_bg.get(ttype, "FFFFFF")

        # Display name with symbol prefix
        disp = name
        if ttype == "Milestone":   disp = "◆ " + name
        elif ttype == "Deliverable": disp = "★ " + name

        vals = {
            "A": (tid,                              "0",          "center"),
            "B": (wbs,                              "@",          "left"),
            "C": (indent,                           "0",          "center"),
            "D": (("  " * indent) + disp,           "@",          "left"),
            "E": (ttype,                            "@",          "left"),
            "F": (status,                           "@",          "left"),
            "G": (priority,                         "@",          "left"),
            "H": (start,                            "YYYY-MM-DD", "center"),
            "I": (end,                              "YYYY-MM-DD", "center"),
            "J": (f"=I{row}-H{row}+1",             "0",          "center"),
            "K": (prog / 100,                       "0%",         "center"),
            "L": (resources[0] if len(resources)>0 else "", "@",  "left"),
            "M": (0.80 if len(resources)>0 else None, "0%",       "center"),
            "N": (resources[1] if len(resources)>1 else "", "@",  "left"),
            "O": (0.20 if len(resources)>1 else None, "0%",       "center"),
            "P": (resources[2] if len(resources)>2 else "", "@",  "left"),
            "Q": (0.20 if len(resources)>2 else None, "0%",       "center"),
            "R": (dur * 600,                        "#,##0",      "right"),
            "S": ("",                               "@",          "left"),
            "T": ("",                               "@",          "left"),
        }

        for ci, (val, fmt, align) in vals.items():
            c = ws[f"{ci}{row}"]
            c.value = val; c.number_format = fmt
            c.fill = fill(bg); c.border = border()
            c.font = Font(size=9, bold=(ttype == "Epic" and ci == "D"),
                          italic=(ttype == "Milestone" and ci == "D"))
            c.alignment = Alignment(horizontal=align, vertical="center")

        # Extra type-based cell styling
        if ttype == "Milestone":
            ws[f"D{row}"].font = Font(bold=True, size=9, color="FF0000")
        elif ttype == "Deliverable":
            ws[f"D{row}"].font = Font(bold=True, size=9, color="7030A0")
        elif ttype == "Epic":
            ws[f"D{row}"].font = Font(bold=True, size=10, color="1F4E79")

        ws.row_dimensions[row].height = 18

    # Conditional formatting — status & priority colours (driven by CONFIG)
    for sname, s_hex in STATUSES:
        ws.conditional_formatting.add(
            f"F4:F{MAX_ROW}",
            FormulaRule(formula=[f'$F4="{sname}"'],
                        fill=fill(s_hex), font=Font(size=9)))
    for pname, p_hex in PRIORITIES:
        ws.conditional_formatting.add(
            f"G4:G{MAX_ROW}",
            FormulaRule(formula=[f'$G4="{pname}"'],
                        fill=fill(p_hex), font=Font(size=9)))
    # Highlight overdue (end < today, not completed)
    ws.conditional_formatting.add(
        f"H4:I{MAX_ROW}",
        FormulaRule(
            formula=[f'AND($I4<TODAY(),$F4<>"Completed",$F4<>"Cancelled")'],
            fill=fill("FFC7CE"), font=Font(size=9, color="9C0006")))

    LAST_ROW = 3 + len(TASKS)
    nr(wb, "TaskIDs",        "TASKS", f"$A$4:$A${LAST_ROW}")
    nr(wb, "TaskNames",      "TASKS", f"$D$4:$D${LAST_ROW}")
    nr(wb, "TaskStartDates", "TASKS", f"$H$4:$H${LAST_ROW}")
    nr(wb, "TaskEndDates",   "TASKS", f"$I$4:$I${LAST_ROW}")
    nr(wb, "TaskProgress",   "TASKS", f"$K$4:$K${LAST_ROW}")
    return ws


# ===========================================================================
# SHEET 4 — GANTT CHART
# ===========================================================================

def build_gantt(wb):
    ws = wb.create_sheet("GANTT_CHART")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "ED7D31"

    # Row layout
    TITLE_ROW  = 1
    LEGEND_ROW = 2
    MONTH_ROW  = 3
    WEEK_ROW   = 4
    DATA_START = 5
    GANTT_COL  = 11   # column K onwards

    # Fixed task-info columns
    INFO_COLS = [
        ("A","ID",   5), ("B","WBS",  8), ("C","Task Name", 34),
        ("D","Type", 11),("E","Status",12),("F","Start",    11),
        ("G","End",  11), ("H","Days",  7), ("I","Done %",   8),
        ("J","Resources",20),
    ]
    for ci, ht, w in INFO_COLS:
        ws.column_dimensions[ci].width = w

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:BJ1")
    ws["A1"].value = '=IFERROR(CFG_ProjectName&"  –  GANTT CHART","PROJECT GANTT CHART")'
    ws["A1"].fill = fill("1F4E79")
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[TITLE_ROW].height = 36

    # ── Legend ───────────────────────────────────────────────────────────────
    col = 1
    for tname, bar_hex, _ in ITEM_TYPES:
        c = ws.cell(row=LEGEND_ROW, column=col)
        c.value = f"  {tname}"
        c.fill = fill(bar_hex)
        c.font = Font(bold=True, size=9,
                      color="FFFFFF" if tname in ("Epic","Milestone") else "000000")
        c.alignment = Alignment(vertical="center")
        c.border = border()
        ws.merge_cells(start_row=LEGEND_ROW, start_column=col,
                       end_row=LEGEND_ROW, end_column=col+1)
        col += 2
    col += 1
    for sname, s_hex in STATUSES:
        c = ws.cell(row=LEGEND_ROW, column=col)
        c.value = sname
        c.fill = fill(s_hex)
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border()
        col += 1
    ws.row_dimensions[LEGEND_ROW].height = 18

    # ── Week header for info columns ─────────────────────────────────────────
    for ci, ht, _ in INFO_COLS:
        c = ws[f"{ci}{WEEK_ROW}"]
        c.value = ht
        hdr(c, "2E75B6", sz=9)
    ws.row_dimensions[WEEK_ROW].height = 28

    # ── Build 52-week timeline ───────────────────────────────────────────────
    # Align to nearest Monday
    d = PROJECT_START
    while d.weekday() != 0:
        d += datetime.timedelta(days=1)

    week_dates = [d + datetime.timedelta(weeks=i) for i in range(52)]

    # Month merged header
    cur_month = None
    mo_start_col = GANTT_COL
    for wi, wd in enumerate(week_dates):
        col = GANTT_COL + wi
        mo = (wd.year, wd.month)
        if mo != cur_month:
            if cur_month is not None:
                end_col = col - 1
                if end_col > mo_start_col:
                    ws.merge_cells(start_row=MONTH_ROW, start_column=mo_start_col,
                                   end_row=MONTH_ROW, end_column=end_col)
                mc = ws.cell(row=MONTH_ROW, column=mo_start_col)
                mc.value = datetime.date(cur_month[0], cur_month[1], 1).strftime("%b %Y")
                mc.fill = fill("2E75B6")
                mc.font = Font(bold=True, size=9, color="FFFFFF")
                mc.alignment = Alignment(horizontal="center", vertical="center")
                mc.border = border()
            cur_month = mo
            mo_start_col = col
        wc = ws.cell(row=WEEK_ROW, column=col)
        wc.value = f"W{wd.strftime('%W')}\n{wd.strftime('%d/%m')}"
        wc.fill = fill("BDD7EE")
        wc.font = Font(bold=True, size=7)
        wc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wc.border = border()
        ws.column_dimensions[get_column_letter(col)].width = 4.2

    # last month
    end_col = GANTT_COL + len(week_dates) - 1
    if end_col > mo_start_col:
        ws.merge_cells(start_row=MONTH_ROW, start_column=mo_start_col,
                       end_row=MONTH_ROW, end_column=end_col)
    mc = ws.cell(row=MONTH_ROW, column=mo_start_col)
    mc.value = datetime.date(cur_month[0], cur_month[1], 1).strftime("%b %Y")
    mc.fill = fill("2E75B6")
    mc.font = Font(bold=True, size=9, color="FFFFFF")
    mc.alignment = Alignment(horizontal="center", vertical="center")
    mc.border = border()
    ws.row_dimensions[MONTH_ROW].height = 18

    # ── Task rows ────────────────────────────────────────────────────────────
    type_bar  = {t[0]: t[1] for t in ITEM_TYPES}
    type_bg   = {t[0]: t[2] for t in ITEM_TYPES}
    today     = datetime.date.today()

    for (tid, wbs, name, ttype, status, priority,
         offset, dur, prog, resources) in TASKS:
        row  = DATA_START + tid - 1
        start = PROJECT_START + datetime.timedelta(days=offset)
        end   = start + datetime.timedelta(days=dur - 1)
        indent = wbs.count(".")
        row_bg = type_bg.get(ttype, "FFFFFF")
        bar_hex = type_bar.get(ttype, "BDD7EE")

        disp = name
        if ttype == "Milestone":   disp = "◆ " + name
        elif ttype == "Deliverable": disp = "★ " + name

        txt_color = ("FF0000" if ttype == "Milestone" else
                     "7030A0" if ttype == "Deliverable" else
                     "1F4E79" if ttype == "Epic" else "000000")
        info = {
            "A": (tid,                            "0",          "center"),
            "B": (wbs,                            "@",          "left"),
            "C": (("  " * indent) + disp,         "@",          "left"),
            "D": (ttype,                          "@",          "left"),
            "E": (status,                         "@",          "left"),
            "F": (start,                          "MM/DD",      "center"),
            "G": (end,                            "MM/DD",      "center"),
            "H": (dur,                            "0",          "center"),
            "I": (prog / 100,                     "0%",         "center"),
            "J": (", ".join(resources[:2]),       "@",          "left"),
        }
        for ci, (val, fmt, align) in info.items():
            c = ws[f"{ci}{row}"]
            c.value = val; c.number_format = fmt
            c.fill = fill(row_bg); c.border = border()
            c.font = Font(size=9, bold=(ttype == "Epic" and ci == "D"),
                          color=txt_color if ci == "D" else "000000")
            c.alignment = Alignment(horizontal=align, vertical="center")

        # Timeline cells
        progress_end = start + datetime.timedelta(days=max(0, int(dur * prog / 100)))
        for wi, wd in enumerate(week_dates):
            col  = GANTT_COL + wi
            c    = ws.cell(row=row, column=col)
            we   = wd + datetime.timedelta(days=6)
            in_task = (wd <= end) and (we >= start)
            c.border = border()

            if in_task:
                if ttype == "Milestone":
                    c.value = "◆"
                    c.fill = fill("FFE0E0")
                    c.font = Font(bold=True, color="FF0000", size=11)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif ttype == "Deliverable":
                    c.value = "★"
                    c.fill = fill("EBD6FF")
                    c.font = Font(bold=True, color="7030A0", size=11)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    if wd <= progress_end:
                        # Completed portion — slightly darker shade
                        r = max(int(bar_hex[0:2], 16) - 45, 0)
                        g = max(int(bar_hex[2:4], 16) - 45, 0)
                        b = max(int(bar_hex[4:6], 16) - 45, 0)
                        darker = f"{r:02X}{g:02X}{b:02X}"
                        c.fill = fill(darker)
                        if prog == 100:
                            c.value = "✓"
                            c.font = Font(size=8, color="FFFFFF", bold=True)
                            c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        c.fill = fill(bar_hex)
            else:
                # Today marker
                if wd <= today <= we:
                    c.fill = fill("FFE0E0")
                else:
                    c.fill = fill("F7F7F7" if wi % 2 else "FFFFFF")

        ws.row_dimensions[row].height = 18

    # Mark today column in header
    for wi, wd in enumerate(week_dates):
        we = wd + datetime.timedelta(days=6)
        if wd <= today <= we:
            tc = ws.cell(row=WEEK_ROW, column=GANTT_COL + wi)
            tc.fill = fill("FF0000")
            tc.font = Font(bold=True, size=7, color="FFFFFF")
            # "TODAY" arrow row 4
            ws.cell(row=MONTH_ROW, column=GANTT_COL + wi).value = "▼ TODAY"
            ws.cell(row=MONTH_ROW, column=GANTT_COL + wi).font = Font(bold=True, size=8, color="FF0000")
            break

    ws.freeze_panes = f"K{DATA_START}"

    # CF: status colours on status column
    LAST = DATA_START + len(TASKS) - 1
    for sname, s_hex in STATUSES:
        ws.conditional_formatting.add(
            f"E{DATA_START}:E{LAST}",
            FormulaRule(formula=[f'$E{DATA_START}="{sname}"'],
                        fill=fill(s_hex)))

    return ws


# ===========================================================================
# SHEET 5 — RESOURCE ALLOCATION
# ===========================================================================

def build_resource_alloc(wb):
    ws = wb.create_sheet("RESOURCE_ALLOC")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "7030A0"
    ws.freeze_panes = "E5"

    ws.merge_cells("A1:Z1")
    ws["A1"].value = "📊  RESOURCE ALLOCATION MATRIX"
    ws["A1"].fill = fill("7030A0")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:Z2")
    ws["A2"].value = (
        "Allocation % per resource per task.  "
        "Column total > 100% = over-allocated (red).  "
        "All resource names sourced from RESOURCES sheet."
    )
    ws["A2"].fill = fill("EAD1FF")
    ws["A2"].font = Font(italic=True, size=9, color="7030A0")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Fixed columns
    for ci, ht, w in [("A","ID",6),("B","WBS",8),("C","Task Name",32),("D","Type",12)]:
        hdr(ws[f"{ci}4"], "7030A0"); ws[f"{ci}4"].value = ht
        ws.column_dimensions[ci].width = w

    # Resource columns
    RES_START_COL = 5
    for i, (rid, name, *_) in enumerate(RESOURCES):
        col = RES_START_COL + i
        cl = get_column_letter(col)
        hdr(ws[f"{cl}4"], "7030A0"); ws[f"{cl}4"].value = name.split()[0]
        ws.column_dimensions[cl].width = 10
        # Resource ID sub-header
        c = ws[f"{cl}5"]
        c.value = rid; c.fill = fill("EAD1FF")
        c.font = Font(size=8, italic=True)
        c.alignment = Alignment(horizontal="center"); c.border = border()

    TOTAL_COL = RES_START_COL + len(RESOURCES)
    tc_l = get_column_letter(TOTAL_COL)
    hdr(ws[f"{tc_l}4"], "7030A0"); ws[f"{tc_l}4"].value = "Task\nTotal"
    ws.column_dimensions[tc_l].width = 10
    ws.row_dimensions[4].height = 22; ws.row_dimensions[5].height = 16

    res_id_to_col = {r[0]: RES_START_COL + i for i, r in enumerate(RESOURCES)}

    DATA_START = 6
    type_bg = {t[0]: t[2] for t in ITEM_TYPES}

    for (tid, wbs, name, ttype, status, priority,
         offset, dur, prog, resources) in TASKS:
        row = DATA_START + tid - 1
        bg = type_bg.get(ttype, "FFFFFF") if tid % 2 else "FFFFFF"

        for ci, val, align in [("A",tid,"center"),("B",wbs,"left"),
                                ("C",name,"left"),("D",ttype,"left")]:
            c = ws[f"{ci}{row}"]
            c.value = val; c.fill = fill(bg)
            c.border = border(); c.font = Font(size=9)
            c.alignment = Alignment(horizontal=align, vertical="center")

        alloc_cells = []
        for i, (rid, *_) in enumerate(RESOURCES):
            col = RES_START_COL + i
            cl  = get_column_letter(col)
            c   = ws[f"{cl}{row}"]
            if rid in resources:
                idx = resources.index(rid)
                alloc = 0.80 if idx == 0 else 0.20
                c.value = alloc
                c.number_format = "0%"
                c.fill = fill("C6EFCE")
                alloc_cells.append(f"{cl}{row}")
            else:
                c.value = None
                c.fill = fill(bg)
            c.border = border(); c.font = Font(size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Task total
        tc = ws[f"{tc_l}{row}"]
        if alloc_cells:
            tc.value = f"=SUM({','.join(alloc_cells)})"
        tc.number_format = "0%"; tc.border = border()
        tc.font = Font(size=9, bold=True)
        tc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 15

    # Totals row
    TOT_ROW = DATA_START + len(TASKS) + 1
    ws[f"C{TOT_ROW}"].value = "TOTAL ALLOCATION"
    ws[f"C{TOT_ROW}"].font = Font(bold=True, size=10)
    ws[f"C{TOT_ROW}"].fill = fill("BDD7EE")

    for i in range(len(RESOURCES)):
        col = RES_START_COL + i
        cl  = get_column_letter(col)
        c   = ws[f"{cl}{TOT_ROW}"]
        c.value = f"=SUM({cl}{DATA_START}:{cl}{TOT_ROW-1})"
        c.number_format = "0%"; c.fill = fill("BDD7EE")
        c.border = border(); c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center")
        # CF: over-allocation
        ws.conditional_formatting.add(
            f"{cl}{TOT_ROW}",
            FormulaRule(formula=[f"{cl}{TOT_ROW}>1"],
                        fill=fill("FFC7CE"), font=Font(bold=True, color="9C0006")))
        ws.conditional_formatting.add(
            f"{cl}{TOT_ROW}",
            FormulaRule(formula=[f"AND({cl}{TOT_ROW}>0,{cl}{TOT_ROW}<=1)"],
                        fill=fill("C6EFCE"), font=Font(bold=True, color="276221")))

    return ws


# ===========================================================================
# SHEET 6 — TIMESHEET
# ===========================================================================

def build_timesheet(wb):
    ws = wb.create_sheet("TIMESHEET")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "C55A11"
    ws.freeze_panes = "E4"

    ws.merge_cells("A1:R1")
    ws["A1"].value = "⏱  MONTHLY TIMESHEET"
    ws["A1"].fill = fill("C55A11")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:R2")
    ws["A2"].value = (
        "Estimated hours per resource per month, derived from allocation % "
        "× task overlap × Working Hours/Day (CONFIG!B11)."
    )
    ws["A2"].fill = fill("FCE4D6")
    ws["A2"].font = Font(italic=True, size=9, color="C55A11")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Build 12-month list
    months = []
    d = PROJECT_START.replace(day=1)
    for _ in range(12):
        months.append(d)
        d = d.replace(month=d.month+1) if d.month < 12 else d.replace(year=d.year+1, month=1)

    FIXED = [("A","Resource",20),("B","Res. ID",9),("C","Task ID",8),("D","Task Name",32)]
    for ci, ht, w in FIXED:
        hdr(ws[f"{ci}3"], "C55A11"); ws[f"{ci}3"].value = ht
        ws.column_dimensions[ci].width = w

    MON_START = 5
    for mi, mdate in enumerate(months):
        col = MON_START + mi
        cl  = get_column_letter(col)
        hdr(ws[f"{cl}3"], "C55A11")
        ws[f"{cl}3"].value = mdate.strftime("%b\n%Y")
        ws.column_dimensions[cl].width = 9

    TOT_COL = MON_START + len(months)
    tc_l    = get_column_letter(TOT_COL)
    hdr(ws[f"{tc_l}3"], "C55A11"); ws[f"{tc_l}3"].value = "Total\nHours"
    ws.column_dimensions[tc_l].width = 10
    ws.row_dimensions[3].height = 28

    # One row per resource per assigned task
    data_row = 4
    res_id_map = {r[0]: r for r in RESOURCES}

    for (tid, wbs, name, ttype, status, priority,
         offset, dur, prog, resources) in TASKS:
        t_start = PROJECT_START + datetime.timedelta(days=offset)
        t_end   = t_start + datetime.timedelta(days=dur - 1)

        for ri, rid in enumerate(resources):
            alloc = 0.80 if ri == 0 else 0.20
            rname = res_id_map.get(rid, (rid, rid))[1]
            bg = "FFF2CC" if data_row % 2 else "FFFFFF"

            for ci, val, align in [("A",rname,"left"),("B",rid,"center"),
                                    ("C",tid,"center"),("D",name,"left")]:
                c = ws[f"{ci}{data_row}"]
                c.value = val; c.fill = fill(bg)
                c.border = border(); c.font = Font(size=9)
                c.alignment = Alignment(horizontal=align, vertical="center")

            month_cells = []
            for mi, mdate in enumerate(months):
                col = MON_START + mi
                cl  = get_column_letter(col)
                c   = ws[f"{cl}{data_row}"]

                # Month range
                if mdate.month == 12:
                    m_end = datetime.date(mdate.year+1, 1, 1) - datetime.timedelta(days=1)
                else:
                    m_end = mdate.replace(month=mdate.month+1) - datetime.timedelta(days=1)

                ov_s = max(t_start, mdate)
                ov_e = min(t_end,   m_end)
                if ov_s <= ov_e:
                    ov_days = (ov_e - ov_s).days + 1
                    hours   = round(ov_days * alloc * WORKING_HOURS_PER_DAY * 5/7, 1)
                    c.value = hours; c.fill = fill("E2EFDA")
                    month_cells.append(f"{cl}{data_row}")
                else:
                    c.value = None; c.fill = fill(bg)
                c.border = border(); c.font = Font(size=9)
                c.alignment = Alignment(horizontal="center")
                c.number_format = "0.0"

            tc = ws[f"{tc_l}{data_row}"]
            tc.value = f"=SUM({','.join(month_cells)})" if month_cells else 0
            tc.number_format = "0.0"
            tc.fill = fill("DEEAF1"); tc.border = border()
            tc.font = Font(bold=True, size=9)
            tc.alignment = Alignment(horizontal="center")
            ws.row_dimensions[data_row].height = 15
            data_row += 1

    # Resource summary totals
    SUM_ROW = data_row + 2
    ws.merge_cells(f"A{SUM_ROW}:D{SUM_ROW}")
    ws[f"A{SUM_ROW}"].value = "TOTAL HOURS BY RESOURCE"
    ws[f"A{SUM_ROW}"].font = Font(bold=True, size=10)
    ws[f"A{SUM_ROW}"].fill = fill("BDD7EE")

    hdr(ws[f"C{SUM_ROW+1}"], "C55A11"); ws[f"C{SUM_ROW+1}"].value = "Resource"
    hdr(ws[f"D{SUM_ROW+1}"], "C55A11"); ws[f"D{SUM_ROW+1}"].value = "Total Hours"

    for i, (rid, rname, *_) in enumerate(RESOURCES):
        r = SUM_ROW + 2 + i
        ws[f"C{r}"].value = rname
        ws[f"D{r}"].value = (
            f'=SUMIF(A4:A{data_row-1},"{rname}",{tc_l}4:{tc_l}{data_row-1})')
        ws[f"D{r}"].number_format = "0.0"
        for ci in ["C","D"]:
            ws[f"{ci}{r}"].border = border()
            ws[f"{ci}{r}"].font = Font(size=10)
        ws[f"D{r}"].font = Font(bold=True, size=10)
        ws[f"D{r}"].alignment = Alignment(horizontal="right")

    return ws


# ===========================================================================
# SHEET 7 — COST ANALYSIS
# ===========================================================================

def build_cost(wb):
    ws = wb.create_sheet("COST_ANALYSIS")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "833C00"
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:N1")
    ws["A1"].value = "💰  COST ANALYSIS"
    ws["A1"].fill = fill("833C00")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    ws["A2"].value = (
        "Costs = resource daily rate (RESOURCES) × allocation % × task duration.  "
        "Budget from TASKS.  Currency from CONFIG."
    )
    ws["A2"].fill = fill("F4B183")
    ws["A2"].font = Font(italic=True, size=9, color="833C00")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Summary box ──────────────────────────────────────────────────────────
    ws.merge_cells("A3:D3")
    ws["A3"].value = "PROJECT COST SUMMARY"
    sec(ws["A3"], "833C00")
    ws.row_dimensions[3].height = 22

    SUMMARY = [
        (4,  "Total Budget",        "=CFG_Budget"),
        (5,  "Contingency Reserve", "=CFG_Budget*CFG_Contingency/100"),
        (6,  "Adjusted Budget",     "=B4+B5"),
        (7,  "Estimated Resource Cost", "=SUM(K13:K500)"),
        (8,  "Estimated Total Cost", "=B7"),
        (9,  "Budget Variance",     "=B6-B8"),
        (10, "Budget Used %",       "=IFERROR(B8/B6,0)"),
    ]
    for r, lab, formula in SUMMARY:
        lc = ws[f"A{r}"]; lc.value = lab; lbl(lc)
        vc = ws[f"B{r}"]; vc.value = formula
        vc.fill = fill("FFFFC0"); vc.border = border()
        vc.font = Font(bold=True, size=11)
        vc.alignment = Alignment(horizontal="right")
        vc.number_format = "#,##0.00" if r <= 9 else "0.00%"

    # Variance CF
    ws.conditional_formatting.add("B9",
        FormulaRule(formula=["$B$9>=0"], fill=fill("C6EFCE"),
                    font=Font(bold=True, color="276221")))
    ws.conditional_formatting.add("B9",
        FormulaRule(formula=["$B$9<0"],  fill=fill("FFC7CE"),
                    font=Font(bold=True, color="9C0006")))

    # ── Per-resource summary ─────────────────────────────────────────────────
    ws.merge_cells("F3:J3")
    ws["F3"].value = "COST BY RESOURCE"
    sec(ws["F3"], "833C00")

    for ci, ht in zip(["F","G","H","I","J"],
                      ["Resource","Role","Hourly Rate","Est. Hours","Est. Cost"]):
        hdr(ws[f"{ci}4"], "C55A11"); ws[f"{ci}4"].value = ht

    for w, ci in [(22,"F"),(18,"G"),(12,"H"),(12,"I"),(14,"J")]:
        ws.column_dimensions[ci].width = w

    for i, (rid, rname, role, dept, avail, rate) in enumerate(RESOURCES, 5):
        bg = "FFF2CC" if i % 2 else "FFFFFF"
        ws[f"F{i}"].value = rname
        ws[f"G{i}"].value = role
        ws[f"H{i}"].value = rate
        ws[f"H{i}"].number_format = "#,##0.00"
        # Hours from timesheet (placeholder – cross-sheet SUMIF)
        ws[f"I{i}"].value = 0
        ws[f"I{i}"].number_format = "0.0"
        ws[f"J{i}"].value = f"=H{i}*I{i}"
        ws[f"J{i}"].number_format = "#,##0.00"
        for ci in ["F","G","H","I","J"]:
            c = ws[f"{ci}{i}"]
            c.border = border(); c.font = Font(size=10); c.fill = fill(bg)
            c.alignment = Alignment(
                horizontal="right" if ci in ("H","I","J") else "left",
                vertical="center")

    # ── Task cost detail ─────────────────────────────────────────────────────
    ws.merge_cells("A12:N12")
    ws["A12"].value = "TASK-LEVEL COST BREAKDOWN"
    sec(ws["A12"], "833C00")

    COST_COLS = [
        ("A","ID",6),("B","WBS",9),("C","Task Name",32),
        ("D","Type",12),("E","Status",12),
        ("F","Start",12),("G","End",12),("H","Days",8),
        ("I","Budget",13),("J","Resource Cost",15),
        ("K","Total Cost",13),("L","Variance",12),
        ("M","Var %",9),("N","Notes",20),
    ]
    for ci, ht, w in COST_COLS:
        hdr(ws[f"{ci}13"], "C55A11"); ws[f"{ci}13"].value = ht
        ws.column_dimensions[ci].width = w

    type_bg = {t[0]: t[2] for t in ITEM_TYPES}
    res_id_map = {r[0]: r for r in RESOURCES}

    for (tid, wbs, name, ttype, status, priority,
         offset, dur, prog, resources) in TASKS:
        row   = 13 + tid
        start = PROJECT_START + datetime.timedelta(days=offset)
        end   = start + datetime.timedelta(days=dur - 1)
        bg    = type_bg.get(ttype, "FFFFFF") if tid % 2 else "FFFFFF"
        budget = dur * 600
        res_cost = sum(
            dur * (0.80 if ri == 0 else 0.20) * res_id_map[rid][5] * WORKING_HOURS_PER_DAY
            for ri, rid in enumerate(resources)
            if rid in res_id_map
        )

        for ci, val, fmt, align in [
            ("A", tid,        "0",          "center"),
            ("B", wbs,        "@",          "left"),
            ("C", name,       "@",          "left"),
            ("D", ttype,      "@",          "left"),
            ("E", status,     "@",          "left"),
            ("F", start,      "YYYY-MM-DD", "center"),
            ("G", end,        "YYYY-MM-DD", "center"),
            ("H", dur,        "0",          "center"),
            ("I", budget,     "#,##0",      "right"),
            ("J", round(res_cost), "#,##0", "right"),
            ("K", f"=I{row}+J{row}", "#,##0","right"),
            ("L", f"=I{row}-K{row}", "#,##0","right"),
            ("M", f'=IFERROR(L{row}/I{row},"")', "0.0%","center"),
            ("N", "",         "@",          "left"),
        ]:
            c = ws[f"{ci}{row}"]
            c.value = val; c.number_format = fmt
            c.fill = fill(bg); c.border = border()
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal=align, vertical="center")
        ws.row_dimensions[row].height = 16

    LAST = 13 + len(TASKS)
    for ci in ["I","J","K","L"]:
        c = ws[f"{ci}{LAST+1}"]
        c.value = f"=SUM({ci}14:{ci}{LAST})"
        c.number_format = "#,##0"; c.fill = fill("BDD7EE")
        c.border = border(); c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="right")
    ws[f"C{LAST+1}"].value = "TOTAL"
    ws[f"C{LAST+1}"].font = Font(bold=True); ws[f"C{LAST+1}"].fill = fill("BDD7EE")

    # Variance CF per task
    for row in range(14, LAST+1):
        ws.conditional_formatting.add(f"L{row}",
            FormulaRule(formula=[f"$L{row}<0"], fill=fill("FFC7CE")))
        ws.conditional_formatting.add(f"L{row}",
            FormulaRule(formula=[f"$L{row}>0"], fill=fill("C6EFCE")))

    return ws


# ===========================================================================
# SHEET 8 — CHANGE LOG
# ===========================================================================

def build_changelog(wb):
    ws = wb.create_sheet("CHANGE_LOG")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "404040"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    ws["A1"].value = "📝  CHANGE LOG"
    ws["A1"].fill = fill("404040")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    COLS = [("A","Date",14),("B","Version",10),("C","Changed By",22),
            ("D","Sheet",16),("E","Cell/Range",14),("F","Old Value",24),
            ("G","New Value",24),("H","Reason / Notes",40)]
    for ci, ht, w in COLS:
        hdr(ws[f"{ci}2"], "404040"); ws[f"{ci}2"].value = ht
        ws.column_dimensions[ci].width = w
    ws.row_dimensions[2].height = 22

    # Seed entry
    for ci, val, fmt in [
        ("A", datetime.date.today(), "YYYY-MM-DD"),
        ("B", "1.0",                 "@"),
        ("C", PROJECT_MANAGER,       "@"),
        ("D", "CONFIG",              "@"),
        ("E", "B5",                  "@"),
        ("F", "",                    "@"),
        ("G", PROJECT_NAME,          "@"),
        ("H", "Initial workbook setup", "@"),
    ]:
        c = ws[f"{ci}3"]; c.value = val; c.number_format = fmt
        c.border = border(); c.font = Font(size=10); c.fill = fill("F2F2F2")
        c.alignment = Alignment(vertical="center")

    return ws


# ===========================================================================
# SHEET 9 — INSTRUCTIONS
# ===========================================================================

def build_instructions(wb):
    ws = wb.create_sheet("INSTRUCTIONS")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "1F4E79"

    ws.merge_cells("A1:E1")
    ws["A1"].value = "ℹ️  HOW TO USE THIS WORKBOOK"
    ws["A1"].fill = fill("1F4E79")
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    for ci, w in [("A",24),("B",52),("C",14),("D",14),("E",14)]:
        ws.column_dimensions[ci].width = w

    ROWS = [
        ("Sheet",              "Purpose / What to edit",                      "Inputs?","Formulas?","Linked to"),
        ("CONFIG",             "Master control panel. Set project name, dates, view type, currency, budget. Define type/status/priority colours. Add holidays. ALL other sheets read from here via named ranges.", "YES – yellow cells", "Partial", "All sheets"),
        ("RESOURCES",          "Add team members, roles, hourly rates, availability. Daily Rate auto-calculates from CONFIG hours/day.", "YES",  "Daily Rate", "TASKS, TIMESHEET, COST"),
        ("TASKS",              "Enter all tasks. Dropdowns for Type/Status/Priority/Resources are driven by CONFIG & RESOURCES named ranges. Duration is a formula.", "YES",  "Duration", "GANTT, ALLOC, TIME, COST"),
        ("GANTT_CHART",        "Read-only visual timeline. 52-week view. Bars coloured by item type. Darker=completed portion. ◆=Milestone ★=Deliverable. Red column=Today.", "No",  "Yes", "TASKS, CONFIG"),
        ("RESOURCE_ALLOC",     "% allocation matrix. Column totals > 100% turn red (over-allocated). All names from RESOURCES.", "Adjust %", "Totals", "RESOURCES, TASKS"),
        ("TIMESHEET",          "Monthly estimated hours per resource per task, based on allocation % and task date overlap.", "Adjust hrs", "Yes", "TASKS, RESOURCES, CONFIG"),
        ("COST_ANALYSIS",      "Task budget vs resource cost breakdown. Green variance = under budget. Red = over. Summary reads budget from CONFIG.", "No", "Yes", "TASKS, RESOURCES, CONFIG"),
        ("CHANGE_LOG",         "Record plan changes. Track version, who changed what cell, old/new values, and reason.", "YES", "No", "—"),
        ("","","","",""),
        ("KEY CONFIG RULES",   "✅ Change colours in CONFIG → they appear in Preview column immediately.",                    "","",""),
        ("",                   "✅ Add rows to RESOURCES → they appear in TASKS resource dropdowns automatically.",           "","",""),
        ("",                   "✅ Gantt View (CONFIG B13): Daily / Weekly / Monthly changes the header description.",        "","",""),
        ("",                   "✅ Currency (CONFIG B12): shown in COST summary label.",                                      "","",""),
        ("",                   "✅ Working Hours/Day (CONFIG B11): drives Daily Rate in RESOURCES.",                          "","",""),
        ("",                   "✅ Holidays (CONFIG Section F): used with NETWORKDAYS formulas.",                             "","",""),
        ("",                   "✅ Budget (CONFIG B14) + Contingency % (B15) → Adjusted Budget in COST.",                    "","",""),
        ("","","","",""),
        ("NAMED RANGES",       "CFG_ProjectName, CFG_StartDate, CFG_EndDate, CFG_HoursPerDay, CFG_Currency, CFG_GanttView, CFG_Budget, CFG_Contingency", "","",""),
        ("",                   "ItemTypeNames, StatusNames, PriorityNames (drive all dropdowns)",                            "","",""),
        ("",                   "ResourceNames, ResourceIDs, ResourceRates, ResourceAvail",                                    "","",""),
        ("",                   "TaskIDs, TaskNames, TaskStartDates, TaskEndDates, TaskProgress",                              "","",""),
    ]

    for ri, row_data in enumerate(ROWS, 2):
        bg = "1F4E79" if ri == 2 else ("EBF3FB" if ri % 2 else "FFFFFF")
        fg = "FFFFFF" if ri == 2 else "000000"
        bold = ri == 2
        for ci_idx, val in enumerate(row_data, 1):
            cl = get_column_letter(ci_idx)
            c  = ws[f"{cl}{ri}"]
            c.value = val; c.fill = fill(bg)
            c.font  = Font(bold=(bold or (ci_idx==1 and val and ri>2 and "KEY" not in val)),
                           size=10, color=fg)
            c.border = border()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if ci_idx == 1 and val and ri > 2:
                c.font = Font(bold=True, size=10, color="1F4E79")
        h = 50 if len(str(row_data[1])) > 90 else 22
        ws.row_dimensions[ri].height = h

    return ws


# ===========================================================================
# MAIN
# ===========================================================================

def main():
    print("Building workbook …")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default "Sheet"

    print("  1/9  CONFIG …")
    build_config(wb)

    print("  2/9  RESOURCES …")
    build_resources(wb)

    print("  3/9  TASKS …")
    build_tasks(wb)

    print("  4/9  GANTT_CHART …")
    build_gantt(wb)

    print("  5/9  RESOURCE_ALLOC …")
    build_resource_alloc(wb)

    print("  6/9  TIMESHEET …")
    build_timesheet(wb)

    print("  7/9  COST_ANALYSIS …")
    build_cost(wb)

    print("  8/9  CHANGE_LOG …")
    build_changelog(wb)

    print("  9/9  INSTRUCTIONS …")
    build_instructions(wb)

    wb.save(OUTPUT_FILE)
    size_kb = os.path.getsize(OUTPUT_FILE) // 1024
    print(f"\n✅  Saved → {OUTPUT_FILE}  ({size_kb} KB)")
    print("    Sheets:", [s.title for s in wb.worksheets])


if __name__ == "__main__":
    main()
