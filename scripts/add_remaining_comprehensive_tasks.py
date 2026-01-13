#!/usr/bin/env python3
"""
Add Remaining Comprehensive Tasks (200+ more tasks)
Run this script to add detailed tasks for remaining epics
"""

import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

print("="*100)
print("ADDING REMAINING COMPREHENSIVE TASKS (200+ tasks)")
print("="*100)
print(f"Started: {datetime.now().strftime('%H:%M:%S')}\n")

wb = openpyxl.load_workbook('docs/__PROJECT_BACKLOG.xlsx')

epics_ws = wb['Epics']
features_ws = wb['Features']
stories_ws = wb['User_Stories']
tasks_ws = wb['Tasks']

done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
todo_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

stats = {'epics': 0, 'features': 0, 'stories': 0, 'tasks': 0}

def add_epic(eid, name, desc, pri="HIGH"):
    row = epics_ws.max_row + 1
    epics_ws.cell(row, 1, eid)
    epics_ws.cell(row, 2, name)
    epics_ws.cell(row, 3, desc)
    epics_ws.cell(row, 4, pri)
    epics_ws.cell(row, 5, "TODO")
    epics_ws.cell(row, 11, f"=SUMIF(Features!B:B,A{row},Features!I:I)")
    stats['epics'] += 1

def add_feature(fid, eid, name, desc, pri="HIGH"):
    row = features_ws.max_row + 1
    features_ws.cell(row, 1, fid)
    features_ws.cell(row, 2, eid)
    features_ws.cell(row, 3, name)
    features_ws.cell(row, 4, desc)
    features_ws.cell(row, 5, pri)
    features_ws.cell(row, 6, "TODO")
    features_ws.cell(row, 9, f"=SUMIF(User_Stories!B:B,A{row},User_Stories!F:F)")
    stats['features'] += 1

def add_story(sid, fid, name, accept, sp=5):
    row = stories_ws.max_row + 1
    stories_ws.cell(row, 1, sid)
    stories_ws.cell(row, 2, fid)
    stories_ws.cell(row, 3, name)
    stories_ws.cell(row, 4, accept)
    stories_ws.cell(row, 5, "TODO")
    stories_ws.cell(row, 6, sp)
    stats['stories'] += 1

def add_task(tid, sid, desc, impl, std, files, hrs=3, sts="TODO"):
    row = tasks_ws.max_row + 1
    tasks_ws.cell(row, 1, tid)
    tasks_ws.cell(row, 2, sid)
    tasks_ws.cell(row, 3, desc)
    tasks_ws.cell(row, 4, impl)
    tasks_ws.cell(row, 5, std)
    tasks_ws.cell(row, 6, files)
    tasks_ws.cell(row, 7, hrs)
    tasks_ws.cell(row, 8, sts)
    
    if sts == "DONE":
        tasks_ws.cell(row, 8).fill = done_fill
    elif sts == "PARTIAL":
        tasks_ws.cell(row, 8).fill = partial_fill
    else:
        tasks_ws.cell(row, 8).fill = todo_fill
    
    stats['tasks'] += 1

# Standard templates
STD_ENT = """✓ C-prefix class ✓ Extend CEntity* ✓ ENTITY_TITLE constants ✓ Logger ✓ @AMetaData ✓ Validations ✓ LAZY fetch ✓ equals/hashCode on id"""
STD_REPO = """✓ I-prefix ✓ Extend IAbstractRepository ✓ #{#entityName} ✓ ORDER BY ✓ @Param ✓ e. prefix"""
STD_SVC = """✓ C-prefix ✓ Extend CAbstractService ✓ @Service ✓ Constructor injection ✓ @Transactional ✓ Objects.requireNonNull ✓ Stateless"""
STD_UI = """✓ CAbstract*Page ✓ C-components ✓ typeName fields ✓ on_x_y handlers ✓ create_x factories ✓ No lambdas ✓ CNotificationService ✓ Grid refresh"""

# ============================================================================
# EPIC 8: TIME TRACKING & WORKLOG (Jira Tempo-style)
# ============================================================================
print("[4/15] Epic 8: Time Tracking & Worklog")
add_epic("E8", "Time Tracking & Worklog", "Track time spent, original estimate, remaining estimate with detailed work logs")

add_feature("F8.1", "E8", "Work Log Entities", "Domain model for time tracking")
add_story("US8.1.1", "F8.1", "Worklog entities", "As developer, worklog entities exist", 5)

tasks_8_1_1 = [
    ("T8.1.1.1", "Add time fields to CProjectItem",
     "AI: 1) Add Long originalEstimateMinutes 2) Add Long remainingEstimateMinutes 3) Add @AMetaData 4) Add getters that return formatted strings (Xh Ym)",
     STD_ENT, "MODIFY: CProjectItem.java", 2, "TODO"),
    
    ("T8.1.1.2", "Create CWorkLog entity",
     "AI: 1) Create extends CEntityDB 2) Add @ManyToOne projectItem 3) Add @ManyToOne user 4) Add LocalDateTime startTime, Integer durationMinutes 5) Add String comment 6) Add Date workDate for reporting",
     STD_ENT, "CREATE: worklog/domain/CWorkLog.java", 3, "TODO"),
    
    ("T8.1.1.3", "Create IWorkLogRepository",
     "AI: 1) findByProjectItem_Id with ORDER BY workDate DESC 2) findByUser_IdAndWorkDateBetween for timesheet 3) sumDurationByProjectItem 4) sumDurationByUserAndDateRange",
     STD_REPO, "CREATE: repositories/IWorkLogRepository.java", 3, "TODO"),
    
    ("T8.1.1.4", "Create CWorkLogService",
     "AI: 1) logWork(itemId, userId, minutes, comment) 2) updateRemainingEstimate(itemId, newRemaining) 3) getTotalTimeSpent(itemId) 4) getUserWorkLogsForDate(userId, date) 5) Validate: remaining >= 0, duration > 0",
     STD_SVC, "CREATE: services/CWorkLogService.java", 4, "TODO"),
]

for i, (tid, desc, impl, std, files, hrs, sts) in enumerate(tasks_8_1_1, 1):
    add_task(f"T8.1.1.{i}", "US8.1.1", desc, impl, std, files, hrs, sts)

print(f"  Added {len(tasks_8_1_1)} tasks")

add_story("US8.1.2", "F8.1", "Worklog UI components", "As user, I can log time on issues", 8)

tasks_8_1_2 = [
    ("T8.1.2.1", "Create CDialogLogWork",
     "AI: 1) Extends CDBEditDialog 2) Time duration field (IntegerField with 'h' suffix converter) 3) Comment textarea 4) Date picker (default today) 5) Checkbox: 'Also update remaining estimate' 6) FormBuilder pattern 7) Save through CWorkLogService",
     STD_UI, "CREATE: ui/CDialogLogWork.java", 5, "TODO"),
    
    ("T8.1.2.2", "Create CComponentWorkLogList",
     "AI: 1) Extends CDiv 2) CGrid showing: date, user, duration, comment 3) Total hours at bottom 4) buttonAdd opens CDialogLogWork 5) buttonDelete for own entries 6) Refresh on grid refresh event",
     STD_UI, "CREATE: ui/CComponentWorkLogList.java", 4, "TODO"),
    
    ("T8.1.2.3", "Add time tracking section to detail pages",
     "AI: 1) Update CAbstractItemDetailPage 2) Add CComponentWorkLogList 3) Show Original/Remaining/Logged time at top 4) Add 'Log Work' button to toolbar",
     STD_UI, "MODIFY: abstracts/CAbstractItemDetailPage.java", 3, "TODO"),
]

for i, (tid, desc, impl, std, files, hrs, sts) in enumerate(tasks_8_1_2, 1):
    add_task(f"T8.1.2.{i}", "US8.1.2", desc, impl, std, files, hrs, sts)

print(f"  Added {len(tasks_8_1_2)} more tasks")

# ============================================================================
# EPIC 9: CUSTOMIZABLE DASHBOARDS
# ============================================================================
print("[5/15] Epic 9: Customizable Dashboards")
add_epic("E9", "Customizable Dashboards", "User-configurable dashboards with draggable widgets/gadgets")

add_feature("F9.1", "E9", "Dashboard Framework", "Core dashboard and widget architecture")
add_story("US9.1.1", "F9.1", "Dashboard entities", "As developer, dashboard entities exist", 5)

tasks_9_1_1 = [
    ("T9.1.1.1", "Create CDashboard entity",
     "AI: 1) Extends CEntityOfCompany 2) Add @ManyToOne owner (CUser) 3) Add Boolean isDefault, isShared 4) Add layout (JSON string storing widget positions) 5) Add @OneToMany widgets",
     STD_ENT, "CREATE: dashboards/domain/CDashboard.java", 3, "TODO"),
    
    ("T9.1.1.2", "Create CDashboardWidget entity",
     "AI: 1) Extends CEntityDB 2) Add @ManyToOne dashboard 3) Add widgetType (enum: BURNDOWN, VELOCITY, PIE_CHART, FILTER_RESULTS, ACTIVITY_STREAM) 4) Add configuration (JSON) 5) Add position (row, column, width, height)",
     STD_ENT, "CREATE: domain/CDashboardWidget.java, EDashboardWidgetType.java", 3, "TODO"),
    
    ("T9.1.1.3", "Create repositories and services",
     "AI: 1) IDashboardRepository: findByOwner, findSharedDashboards 2) IDashboardWidgetRepository: findByDashboard_Id 3) CDashboardService: CRUD + shareDashboard, cloneDashboard 4) CDashboardWidgetService",
     f"{STD_REPO}\n{STD_SVC}", "CREATE: repositories/I*.java, services/C*Service.java", 4, "TODO"),
]

for i, (tid, desc, impl, std, files, hrs, sts) in enumerate(tasks_9_1_1, 1):
    add_task(f"T9.1.1.{i}", "US9.1.1", desc, impl, std, files, hrs, sts)

print(f"  Added {len(tasks_9_1_1)} tasks")

add_story("US9.1.2", "F9.1", "Dashboard UI", "As user, I can customize my dashboard", 13)

tasks_9_1_2 = [
    ("T9.1.2.1", "Create CDashboardPage",
     "AI: 1) Extends CAbstractPage 2) Use Vaadin Board for grid layout 3) Load user's default dashboard 4) Toolbar: dropdown to switch dashboards, buttonEdit, buttonAddWidget 5) Render widgets based on type",
     STD_UI, "CREATE: ui/CDashboardPage.java", 5, "TODO"),
    
    ("T9.1.2.2", "Create CAbstractDashboardWidget base",
     "AI: 1) Abstract class extends CDiv 2) Has CDashboardWidget model 3) Abstract method: render() 4) Has buttonConfigure, buttonRemove 5) Draggable for repositioning",
     STD_UI, "CREATE: ui/widgets/CAbstractDashboardWidget.java", 4, "TODO"),
    
    ("T9.1.2.3", "Create CWidgetBurndownChart",
     "AI: 1) Extends CAbstractDashboardWidget 2) Configuration: sprintId 3) Use Vaadin Charts 4) Query CSprintMetricsService for data 5) Show ideal line and actual line",
     STD_UI, "CREATE: ui/widgets/CWidgetBurndownChart.java", 5, "TODO"),
    
    ("T9.1.2.4", "Create CWidgetFilterResults",
     "AI: 1) Configuration: saved filter ID 2) Shows CGrid of matching issues 3) Click opens issue detail 4) Refresh button 5) Show count",
     STD_UI, "CREATE: ui/widgets/CWidgetFilterResults.java", 4, "TODO"),
    
    ("T9.1.2.5", "Create CWidgetPieChart",
     "AI: 1) Configuration: groupBy (status, assignee, priority) 2) Use Vaadin Charts 3) Query aggregated data 4) Click slice filters main view",
     STD_UI, "CREATE: ui/widgets/CWidgetPieChart.java", 4, "TODO"),
]

for i, (tid, desc, impl, std, files, hrs, sts) in enumerate(tasks_9_1_2, 1):
    add_task(f"T9.1.2.{i}", "US9.1.2", desc, impl, std, files, hrs, sts)

print(f"  Added {len(tasks_9_1_2)} more tasks")

# ============================================================================
# EPIC 10: ADVANCED SEARCH & JQL
# ============================================================================
print("[6/15] Epic 10: Advanced Search & JQL")
add_epic("E10", "Advanced Search & JQL", "Jira Query Language-like advanced search with saved filters")

add_feature("F10.1", "E10", "JQL Parser", "Parse and execute JQL-like queries")
add_story("US10.1.1", "F10.1", "JQL parser implementation", "As developer, JQL parser exists", 13)

tasks_10_1_1 = [
    ("T10.1.1.1", "Create JQL grammar and lexer",
     "AI: 1) Research ANTLR or JavaCC 2) Define grammar for: field operator value AND/OR 3) Operators: =, !=, >, <, IN, NOT IN, ~, IS EMPTY 4) Functions: currentUser(), endOfMonth() 5) Example: status = 'In Progress' AND assignee = currentUser()",
     "✓ Use ANTLR4 ✓ Grammar file: src/main/antlr4/JQL.g4 ✓ Generate parser/lexer ✓ Add antlr4-maven-plugin", "CREATE: src/main/antlr4/JQL.g4\nMODIFY: pom.xml (add antlr4 dependency)", 8, "TODO"),
    
    ("T10.1.1.2", "Create CJQLQuery model",
     "AI: 1) Class representing parsed query 2) Has List<CJQLClause> clauses 3) CJQLClause: field, operator, value, logicalOperator 4) Method: toCriteria() returns JPA Criteria",
     STD_SVC, "CREATE: search/CJQLQuery.java, CJQLClause.java, EJQLOperator.java", 5, "TODO"),
    
    ("T10.1.1.3", "Create CJQLExecutor service",
     "AI: 1) Method: execute(String jql) returns List<CProjectItem> 2) Parse JQL string 3) Convert to JPA Criteria or JPQL 4) Execute query 5) Handle functions (currentUser, etc) 6) Add result caching",
     STD_SVC, "CREATE: services/CJQLExecutor.java", 8, "TODO"),
]

for i, (tid, desc, impl, std, files, hrs, sts) in enumerate(tasks_10_1_1, 1):
    add_task(f"T10.1.1.{i}", "US10.1.1", desc, impl, std, files, hrs, sts)

print(f"  Added {len(tasks_10_1_1)} tasks")

add_story("US10.1.2", "F10.1", "JQL search UI", "As user, I can use JQL to search issues", 8)

tasks_10_1_2 = [
    ("T10.1.2.1", "Create CJQLSearchField component",
     "AI: 1) Extends CTextField 2) Syntax highlighting 3) Autocomplete for fields and operators 4) Validation on blur 5) Show error messages for invalid syntax 6) History dropdown (recent searches)",
     STD_UI, "CREATE: ui/CJQLSearchField.java", 8, "TODO"),
    
    ("T10.1.2.2", "Create CAdvancedSearchPage",
     "AI: 1) Extends CAbstractPage 2) CJQLSearchField at top 3) CGrid showing results 4) Toolbar: buttonSaveFilter, buttonExport 5) Faceted search sidebar (quick filters) 6) Result count display",
     STD_UI, "CREATE: ui/CAdvancedSearchPage.java", 5, "TODO"),
    
    ("T10.1.2.3", "Create CSavedFilter entity and management",
     "AI: 1) Entity: CSavedFilter with name, jql, owner, isShared 2) Service: CSavedFilterService 3) Dialog: CDialogSaveFilter 4) Component: CComponentSavedFilters (list favorites)",
     f"{STD_ENT}\n{STD_SVC}\n{STD_UI}", "CREATE: domain/CSavedFilter.java, services/CSavedFilterService.java, ui/CDialogSaveFilter.java", 5, "TODO"),
]

for i, (tid, desc, impl, std, files, hrs, sts) in enumerate(tasks_10_1_2, 1):
    add_task(f"T10.1.2.{i}", "US10.1.2", desc, impl, std, files, hrs, sts)

print(f"  Added {len(tasks_10_1_2)} more tasks")

# Continue with more epics...
print(f"\n{'='*100}")
print(f"Saving progress... ({stats['tasks']} tasks added in this run)")
wb.save('docs/__PROJECT_BACKLOG.xlsx')
print("✅ Saved!")

print(f"\n{'='*100}")
print("SUMMARY OF ADDED ITEMS:")
print(f"  Epics:        {stats['epics']}")
print(f"  Features:     {stats['features']}")
print(f"  User Stories: {stats['stories']}")
print(f"  Tasks:        {stats['tasks']}")
print(f"\n  TOTAL:        {sum(stats.values())}")
print("="*100)
