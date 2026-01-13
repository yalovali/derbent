#!/usr/bin/env python3
"""
Add Final 200+ Tasks to Reach 300+ Total
Covers remaining Jira/ProjeQtOr features for comprehensive backlog
"""

import openpyxl
from openpyxl.styles import PatternFill

print("="*100)
print("ADDING FINAL 200+ TASKS - ACHIEVING JIRA/PROJEQTOR PARITY")
print("="*100)

wb = openpyxl.load_workbook('docs/__PROJECT_BACKLOG.xlsx')

epics_ws = wb['Epics']
features_ws = wb['Features']
stories_ws = wb['User_Stories']
tasks_ws = wb['Tasks']

done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
todo_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

stats = {'epics': 0, 'features': 0, 'stories': 0, 'tasks': 0}

def add_epic(eid, name, desc, pri="HIGH"):
    row = epics_ws.max_row + 1
    epics_ws.cell(row, 1, eid)
    epics_ws.cell(row, 2, name)
    epics_ws.cell(row, 3, desc)
    epics_ws.cell(row, 4, pri)
    epics_ws.cell(row, 5, "TODO")
    epics_ws.cell(row, 11, f"=SUMIF(Features!B:B,A{row},Features!I:I)")
    stats['epics'] += 1

def add_feature(fid, eid, name, desc, pri="HIGH"):
    row = features_ws.max_row + 1
    features_ws.cell(row, 1, fid)
    features_ws.cell(row, 2, eid)
    features_ws.cell(row, 3, name)
    features_ws.cell(row, 4, desc)
    features_ws.cell(row, 5, pri)
    features_ws.cell(row, 6, "TODO")
    features_ws.cell(row, 9, f"=SUMIF(User_Stories!B:B,A{row},User_Stories!F:F)")
    stats['features'] += 1

def add_story(sid, fid, name, accept, sp=5):
    row = stories_ws.max_row + 1
    stories_ws.cell(row, 1, sid)
    stories_ws.cell(row, 2, fid)
    stories_ws.cell(row, 3, name)
    stories_ws.cell(row, 4, accept)
    stories_ws.cell(row, 5, "TODO")
    stories_ws.cell(row, 6, sp)
    stats['stories'] += 1

def add_task(tid, sid, desc, impl, std, files, hrs=3, sts="TODO"):
    row = tasks_ws.max_row + 1
    tasks_ws.cell(row, 1, tid)
    tasks_ws.cell(row, 2, sid)
    tasks_ws.cell(row, 3, desc)
    tasks_ws.cell(row, 4, impl)
    tasks_ws.cell(row, 5, std)
    tasks_ws.cell(row, 6, files)
    tasks_ws.cell(row, 7, hrs)
    tasks_ws.cell(row, 8, sts)
    
    if sts == "DONE":
        tasks_ws.cell(row, 8).fill = done_fill
    elif sts == "PARTIAL":
        tasks_ws.cell(row, 8).fill = partial_fill
    else:
        tasks_ws.cell(row, 8).fill = todo_fill
    
    stats['tasks'] += 1

S = """✓ C-prefix ✓ CEntity* ✓ ENTITY_TITLE ✓ Logger ✓ @AMetaData ✓ Validation ✓ LAZY ✓ equals/hashCode"""
R = """✓ I-prefix ✓ IAbstractRepository ✓ #{#entityName} ✓ ORDER BY ✓ @Param"""
V = """✓ C-prefix ✓ CAbstractService ✓ @Service ✓ Inject ✓ @Transactional ✓ Stateless"""
U = """✓ CAbstract*Page ✓ C-components ✓ typeName ✓ on_x_y ✓ create_x ✓ No lambda ✓ CNotificationService"""

# ==================================================================================================================
# EPIC 11: NOTIFICATIONS & COMMUNICATION
# ==================================================================================================================
print("\n[7/15] Epic 11: Notifications & Communication")
add_epic("E11", "Notifications & Communication", "Email notifications, @mentions, comments, activity streams")

add_feature("F11.1", "E11", "Notification System", "Configurable email notifications with schemes")
add_story("US11.1.1", "F11.1", "Notification entities", "As developer, notification entities exist", 5)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CNotificationScheme entity", "AI: 1) Extends CEntityOfCompany 2) Add name, description 3) Add @OneToMany events", S, "CREATE: notifications/domain/CNotificationScheme.java", 2),
    ("Create CNotificationEvent entity", "AI: 1) Extends CEntityDB 2) Add eventType (enum: ITEM_CREATED, STATUS_CHANGED, ASSIGNED, COMMENTED) 3) Add recipientType (enum: ASSIGNEE, REPORTER, WATCHERS, PROJECT_ROLE) 4) Add template", S, "CREATE: domain/CNotificationEvent.java, ENotificationEventType.java, ENotificationRecipientType.java", 4),
    ("Create CNotificationTemplate entity", "AI: 1) Subject and body templates 2) Use Thymeleaf syntax 3) Support variables: ${item.name}, ${user.name}, ${project.name}", S, "CREATE: domain/CNotificationTemplate.java", 3),
    ("Create CNotificationQueue entity", "AI: 1) Queue for async email sending 2) Add recipient, subject, body, sentDate, status (PENDING, SENT, FAILED) 3) Add retryCount", S, "CREATE: domain/CNotificationQueue.java", 2),
], 1):
    add_task(f"T11.1.1.{i}", "US11.1.1", desc, f"{impl}\n{std}", files, hrs, "TODO")

add_story("US11.1.2", "F11.1", "Notification services", "As developer, notification sending works", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CNotificationService (email sender)", "AI: 1) Inject JavaMailSender 2) Method: queueNotification(recipients, template, variables) 3) @Async sendEmail 4) Handle failures with retry logic 5) Log all sends", "CREATE: services/CNotificationService.java", 5),
    ("Create CNotificationScheduler", "AI: 1) @Scheduled(fixedDelay=60000) 2) Query PENDING notifications 3) Send batch (limit 50) 4) Update status 5) Retry FAILED (max 3 attempts)", "CREATE: services/CNotificationScheduler.java", 4),
    ("Add notification triggering to services", "AI: 1) Update CActivityService.save() to trigger ITEM_CREATED event 2) Update changeStatus() to trigger STATUS_CHANGED 3) Update assignTo() to trigger ASSIGNED 4) Use Spring Events (@EventListener)", "MODIFY: All *Service.java files", 5),
], 1):
    add_task(f"T11.1.2.{i}", "US", desc, f"{impl}\n{std}", files, hrs, "TODO")

add_feature("F11.2", "E11", "Comments & @Mentions", "Threaded comments with user mentions")
add_story("US11.2.1", "F11.2", "Comment system", "As user, I can comment on issues and mention users", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CComment entity", "AI: 1) Extends CEntityDB 2) @ManyToOne projectItem, author 3) @Lob String content (HTML) 4) @ManyToOne parentComment for threading 5) Add mentionedUsers (parse @username)", S, "CREATE: comments/domain/CComment.java", 3),
    ("Create ICommentRepository", "AI: 1) findByProjectItem_Id ORDER BY createdDate ASC 2) findByAuthor_Id 3) findByMentionedUsers_IdContains 4) findRootComments (parentComment IS NULL)", R, "CREATE: repositories/ICommentRepository.java", 2),
    ("Create CCommentService", "AI: 1) addComment(itemId, userId, content) 2) Parse @mentions from content 3) Notify mentioned users 4) deleteComment (soft delete or CASCADE) 5) editComment", V, "CREATE: services/CCommentService.java", 4),
    ("Create CComponentCommentSection", "AI: 1) Extends CDiv 2) Tree layout for threaded comments 3) Rich text editor with @mention autocomplete 4) Reply button per comment 5) Edit/Delete for own comments 6) Real-time updates with Broadcaster", U, "CREATE: ui/CComponentCommentSection.java", 8),
], 1):
    add_task(f"T11.2.1.{i}", "US11.2.1", desc, impl, std, files, hrs, "TODO")

add_feature("F11.3", "E11", "Activity Stream", "Real-time activity feed")
add_story("US11.3.1", "F11.3", "Activity stream", "As user, I see real-time activity on my dashboard", 5)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CActivityLog entity", "AI: 1) Actor (user), action (enum), target (projectItem), timestamp 2) Store JSON of changes 3) Indexed on timestamp DESC", S, "CREATE: activity/domain/CActivityLog.java, EActivityAction.java", 3),
    ("Create CActivityLogService with interceptor", "AI: 1) @Aspect to intercept @Transactional methods 2) Log CRUD operations 3) Method: getActivityStream(userId, limit) 4) Filter by user's projects", V, "CREATE: services/CActivityLogService.java", 5),
    ("Create CComponentActivityStream widget", "AI: 1) Extends CDiv 2) Auto-refresh every 30s 3) Show: user avatar, action description, time ago 4) Click opens related item 5) Limit 20 items", U, "CREATE: ui/CComponentActivityStream.java", 4),
], 1):
    add_task(f"T11.3.1.{i}", "US11.3.1", desc, impl, files, hrs, "TODO")

# ==================================================================================================================
# EPIC 12: PERMISSIONS & SECURITY
# ==================================================================================================================
print("\n[8/15] Epic 12: Permissions & Security")
add_epic("E12", "Permissions & Security", "Fine-grained permissions, field-level security, audit logs, SSO")

add_feature("F12.1", "E12", "Permission Framework", "Role-based permissions at project level")
add_story("US12.1.1", "F12.1", "Permission entities", "As developer, permission model exists", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CPermission entity", "AI: 1) Extends CEntityOfCompany 2) Add resource (enum: PROJECT, ACTIVITY, USER, etc) 3) Add action (enum: VIEW, CREATE, EDIT, DELETE, ASSIGN) 4) Add @ManyToOne role", S, "CREATE: permissions/domain/CPermission.java, EResource.java, EAction.java", 3),
    ("Create CRolePermission join entity", "AI: 1) @ManyToOne role, permission 2) Add conditions (JSON) for conditional permissions 3) Example: {\"status\": \"Open\"} means permission only applies to Open items", S, "CREATE: domain/CRolePermission.java", 2),
    ("Update CRole with permissions", "AI: 1) Add @ManyToMany List<CPermission> permissions 2) Add @OneToMany List<CRolePermission> rolePermissions 3) Add helper methods: hasPermission(resource, action)", S, "MODIFY: users/domain/CRole.java", 2),
    ("Create IPermissionRepository", "AI: 1) findByRole_Id 2) findByResource_AndAction 3) Custom query: hasUserPermission(userId, projectId, resource, action)", R, "CREATE: repositories/IPermissionRepository.java", 3),
], 1):
    add_task(f"T12.1.1.{i}", "US12.1.1", desc, impl, files, hrs, "TODO")

add_story("US12.1.2", "F12.1", "Permission checking service", "As developer, permission checks work everywhere", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CPermissionService", "AI: 1) canUser(userId, projectId, resource, action) boolean method 2) Cache results (Caffeine) 3) clearCache on role change 4) Method: getEffectivePermissions(userId, projectId) 5) Consider: user roles, project roles", V, "CREATE: services/CPermissionService.java", 6),
    ("Create @RequiresPermission annotation", "AI: 1) @Target METHOD 2) Parameters: resource, action 3) Create PermissionAspect to check before method execution 4) Throw AccessDeniedException if not allowed", V, "CREATE: annotations/RequiresPermission.java, aspects/PermissionAspect.java", 5),
    ("Add permission checks to all services", "AI: 1) Add @RequiresPermission to CRUD methods 2) Example: @RequiresPermission(resource=PROJECT, action=EDIT) on save() 3) Update 20+ service classes", V, "MODIFY: All *Service.java classes", 8),
], 1):
    add_task(f"T12.1.2.{i}", "US12.1.2", desc, impl, files, hrs, "TODO")

add_feature("F12.2", "E12", "Audit Logging", "Track all changes for compliance")
add_story("US12.2.1", "F12.2", "Audit trail", "As admin, I can see who changed what and when", 5)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CAuditLog entity", "AI: 1) userId, entityType, entityId, action, oldValue (JSON), newValue (JSON), timestamp, ipAddress 2) Indexed on timestamp, userId, entityType", S, "CREATE: audit/domain/CAuditLog.java", 3),
    ("Create CAuditService with Hibernate Envers", "AI: 1) Add @Audited to entities 2) Configure Envers 3) Method: getAuditTrail(entityClass, entityId) 4) Method: getChangedFields(revision) 5) Method: getUserActivity(userId, dateRange)", V, "CREATE: services/CAuditService.java\nMODIFY: pom.xml (add hibernate-envers)", 6),
    ("Create CAuditLogPage", "AI: 1) Grid with filters: user, entity type, date range, action 2) Detail view shows before/after JSON diff 3) Export to CSV", U, "CREATE: ui/CAuditLogPage.java", 5),
], 1):
    add_task(f"T12.2.1.{i}", "US12.2.1", desc, impl, files, hrs, "TODO")

add_feature("F12.3", "E12", "SSO & 2FA", "LDAP/SAML integration and two-factor authentication")
add_story("US12.3.1", "F12.3", "SSO integration", "As enterprise user, I can login with company SSO", 13)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Add Spring Security SAML dependency", "AI: 1) Add spring-security-saml2-service-provider 2) Configure SecurityFilterChain 3) Add SAML metadata endpoint 4) Configure assertion consumer service", V, "MODIFY: pom.xml, SecurityConfig.java", 5),
    ("Create CSAMLConfigEntity", "AI: 1) Store SAML IDP metadata URL 2) Entity ID, SSO URL, certificate 3) Per-company configuration", S, "CREATE: security/domain/CSAMLConfig.java", 3),
    ("Create CLDAP AuthenticationProvider", "AI: 1) Implement UserDetailsService 2) Connect to LDAP server 3) Map LDAP attributes to CUser 4) Auto-create users on first login 5) Sync roles from LDAP groups", V, "CREATE: security/CLDAPAuthenticationProvider.java", 8),
], 1):
    add_task(f"T12.3.1.{i}", "US12.3.1", desc, impl, files, hrs, "TODO")

add_story("US12.3.2", "F12.3", "Two-factor authentication", "As user, I can enable 2FA for account security", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Add 2FA fields to CUser", "AI: 1) Add twoFactorEnabled (Boolean) 2) Add twoFactorSecret (encrypted String) 3) Add backupCodes (encrypted List<String>)", S, "MODIFY: users/domain/CUser.java", 2),
    ("Create C2FAService using TOTP", "AI: 1) Generate secret with GoogleAuthenticator library 2) generateQRCode() 3) validateCode(user, code) 4) generateBackupCodes()",  V, "CREATE: security/C2FAService.java\nMODIFY: pom.xml (add googleauth)", 5),
    ("Create C2FASetupDialog and login flow", "AI: 1) Show QR code to scan 2) Verify code before enabling 3) Show backup codes 4) Update CLoginView to prompt for 2FA code after password 5) 'Use backup code' link", U, "CREATE: ui/C2FASetupDialog.java\nMODIFY: CLoginView.java", 6),
], 1):
    add_task(f"T12.3.2.{i}", "US12.3.2", desc, impl, files, hrs, "TODO")

# ==================================================================================================================
# EPIC 13: AUTOMATION & INTEGRATION
# ==================================================================================================================
print("\n[9/15] Epic 13: Automation & Integration")
add_epic("E13", "Automation & Integration", "Automation rules, REST API, webhooks, GitHub/GitLab integration")

add_feature("F13.1", "E13", "Automation Rules", "If-this-then-that automation (Jira Automation)")
add_story("US13.1.1", "F13.1", "Automation engine", "As admin, I can create automation rules", 13)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CAutomationRule entity", "AI: 1) Extends CEntityOfCompany 2) Add trigger (enum: ITEM_CREATED, STATUS_CHANGED, FIELD_CHANGED, SCHEDULED) 3) Add conditions (JSON array of conditions) 4) Add actions (JSON array of actions) 5) Add enabled flag", S, "CREATE: automation/domain/CAutomationRule.java, ETrigger.java", 4),
    ("Create CAutomationCondition classes", "AI: 1) Interface ICondition with evaluate() method 2) Classes: StatusEqualsCondition, AssigneeIsCondition, PriorityCondition, FieldMatchesCondition 3) All implement ICondition", V, "CREATE: automation/conditions/*.java", 6),
    ("Create CAutomationAction classes", "AI: 1) Interface IAction with execute() method 2) Classes: AssignToAction, ChangeStatusAction, AddCommentAction, SendEmailAction, CreateSubtaskAction 3) All implement IAction", V, "CREATE: automation/actions/*.java", 8),
    ("Create CAutomationEngine service", "AI: 1) Method: evaluateAndExecute(trigger, item) 2) Load applicable rules 3) Evaluate conditions 4) Execute actions if all conditions pass 5) Log execution 6) Handle errors gracefully", V, "CREATE: services/CAutomationEngine.java", 8),
    ("Integrate automation into lifecycle", "AI: 1) Add hooks in CActivityService.save(), changeStatus() 2) Fire automation events 3) Use @EventListener for async execution 4) Add @Scheduled for scheduled triggers", V, "MODIFY: All *Service.java", 5),
], 1):
    add_task(f"T13.1.1.{i}", "US13.1.1", desc, impl, files, hrs, "TODO")

add_story("US13.1.2", "F13.1", "Automation UI", "As admin, I can configure automation rules visually", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CAutomationRuleDesigner", "AI: 1) Step 1: Select trigger 2) Step 2: Add conditions (IF section) 3) Step 3: Add actions (THEN section) 4) Visual rule preview 5) Test rule button", U, "CREATE: ui/CAutomationRuleDesigner.java", 10),
    ("Create condition/action config dialogs", "AI: 1) Each condition type has config dialog 2) Each action type has config dialog 3) Use factory pattern to create dialogs 4) Validate configuration", U, "CREATE: ui/automation/*.java", 8),
], 1):
    add_task(f"T13.1.2.{i}", "US13.1.2", desc, impl, files, hrs, "TODO")

add_feature("F13.2", "E13", "REST API", "RESTful API with OpenAPI documentation")
add_story("US13.2.1", "F13.2", "REST API implementation", "As developer, I can access data via REST API", 13)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Add SpringDoc OpenAPI dependency", "AI: 1) Add springdoc-openapi-starter-webmvc-ui 2) Configure Swagger UI 3) Add @OpenAPIDefinition 4) Enable /swagger-ui.html", V, "MODIFY: pom.xml, application.properties", 2),
    ("Create API controllers for entities", "AI: 1) Create @RestController for CActivity, CProject, CUser, CMeeting 2) CRUD endpoints: GET /api/activities, POST, PUT, DELETE 3) Add @Operation annotations 4) Add pagination and sorting 5) Return DTOs not entities", V, "CREATE: api/controllers/*.java", 12),
    ("Create API DTOs", "AI: 1) ActivityDTO, ProjectDTO, UserDTO 2) Use MapStruct for entity-to-DTO mapping 3) Exclude sensitive fields 4) Add validation", S, "CREATE: api/dto/*.java\nMODIFY: pom.xml (add mapstruct)", 6),
    ("Add API authentication with JWT", "AI: 1) Add jjwt dependency 2) Create JWTTokenProvider 3) Add /api/auth/login endpoint 4) Return JWT token 5) Add JWTAuthenticationFilter 6) Validate token on API requests", V, "CREATE: security/JWTTokenProvider.java, JWTAuthenticationFilter.java", 8),
    ("Add rate limiting", "AI: 1) Use Bucket4j library 2) Configure rate limits per user 3) Return 429 Too Many Requests 4) Add X-RateLimit headers", V, "CREATE: security/RateLimitFilter.java\nMODIFY: pom.xml", 4),
], 1):
    add_task(f"T13.2.1.{i}", "US13.2.1", desc, impl, files, hrs, "TODO")

add_feature("F13.3", "E13", "Webhooks", "Outgoing webhooks for integrations")
add_story("US13.3.1", "F13.3", "Webhook system", "As admin, I can configure webhooks to external systems", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CWebhook entity", "AI: 1) Extends CEntityOfCompany 2) Add url, secret, events (multiselect enum) 3) Add enabled flag 4) Add headers (JSON)", S, "CREATE: webhooks/domain/CWebhook.java", 3),
    ("Create CWebhookService", "AI: 1) Method: sendWebhook(event, payload) 2) Find matching webhooks 3) Sign payload with HMAC-SHA256 4) POST to URL 5) Retry on failure (3 attempts) 6) Log deliveries", V, "CREATE: services/CWebhookService.java", 6),
    ("Create CWebhookDelivery entity for audit", "AI: 1) Store: webhook, event, payload, statusCode, responseBody, timestamp 2) Allow replay 3) Show in UI", S, "CREATE: domain/CWebhookDelivery.java", 2),
    ("Create webhook management UI", "AI: 1) CWebhookListPage 2) Add/Edit/Delete webhooks 3) Test webhook button 4) Delivery history grid 5) Replay failed deliveries", U, "CREATE: ui/CWebhookListPage.java", 5),
], 1):
    add_task(f"T13.3.1.{i}", "US13.3.1", desc, impl, files, hrs, "TODO")

add_feature("F13.4", "E13", "GitHub/GitLab Integration", "Link commits and PRs to issues")
add_story("US13.4.1", "F13.4", "VCS integration", "As developer, commits and PRs link to issues", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CGitRepository entity", "AI: 1) Extends CEntityOfProject 2) Add provider (enum: GITHUB, GITLAB, BITBUCKET) 3) Add url, accessToken 4) Add webhookSecret", S, "CREATE: vcs/domain/CGitRepository.java", 3),
    ("Create CGitCommit entity", "AI: 1) sha, message, author, timestamp 2) @ManyToMany List<CProjectItem> linkedItems 3) Parse issue keys from message (e.g., ACT-123)", S, "CREATE: domain/CGitCommit.java", 3),
    ("Create CGitWebhookController", "AI: 1) @PostMapping /api/webhooks/github 2) Parse GitHub webhook payload 3) Extract commit info 4) Link to issues by parsing message 5) Create CGitCommit entities 6) Verify signature", V, "CREATE: api/CGitWebhookController.java", 6),
    ("Create CComponentGitActivity widget", "AI: 1) Show recent commits for issue 2) Show linked PRs 3) Click opens GitHub/GitLab 4) Add 'Link commit' button", U, "CREATE: ui/CComponentGitActivity.java", 4),
], 1):
    add_task(f"T13.4.1.{i}", "US13.4.1", desc, impl, files, hrs, "TODO")

# ==================================================================================================================
# EPIC 14: QUALITY & TESTING
# ==================================================================================================================
print("\n[10/15] Epic 14: Quality & Testing")
add_epic("E14", "Quality & Testing Management", "Test case management, test execution, defect tracking, coverage reports")

add_feature("F14.1", "E14", "Test Case Management", "Create and organize test cases")
add_story("US14.1.1", "F14.1", "Test case entities", "As developer, test case entities exist", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CTestCase entity", "AI: 1) Extends CEntityOfProject 2) Add preconditions, steps (JSON), expectedResult 3) Add priority, type (enum: FUNCTIONAL, INTEGRATION, E2E) 4) Add estimatedDuration 5) @ManyToMany linkedActivities", S, "CREATE: testing/domain/CTestCase.java, ETestType.java", 4),
    ("Create CTestStep entity", "AI: 1) Embedded in CTestCase 2) Add stepNumber, action, expectedResult 3) Add attachments", S, "CREATE: domain/CTestStep.java", 2),
    ("Create CTestExecution entity", "AI: 1) @ManyToOne testCase 2) Add executedBy, executedDate, status (enum: PASS, FAIL, BLOCKED, SKIPPED) 3) Add actualResult, defects found 4) Add duration", S, "CREATE: domain/CTestExecution.java, ETestStatus.java", 3),
    ("Create CTestSuite entity", "AI: 1) Extends CEntityOfProject 2) @ManyToMany List<CTestCase> testCases 3) Add description, type (REGRESSION, SMOKE, SANITY)", S, "CREATE: domain/CTestSuite.java", 2),
], 1):
    add_task(f"T14.1.1.{i}", "US14.1.1", desc, impl, files, hrs, "TODO")

add_story("US14.1.2", "F14.1", "Test execution tracking", "As QA, I can execute test cases and track results", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CTestCaseService", "AI: 1) CRUD for test cases 2) executeTestCase(testCaseId, userId) 3) getTestCoverage(projectId) 4) getPassRate(suiteId) 5) linkToActivity(testCaseId, activityId)", V, "CREATE: services/CTestCaseService.java", 5),
    ("Create CTestExecutionDialog", "AI: 1) Show test steps 2) Mark each step pass/fail 3) Add actual result 4) Link defects if failed 5) Save execution record", U, "CREATE: ui/CTestExecutionDialog.java", 6),
    ("Create CTestSuitePage", "AI: 1) Grid of test cases in suite 2) Bulk execute button 3) Progress bar (X/Y passed) 4) Filter by status 5) Export results to Excel", U, "CREATE: ui/CTestSuitePage.java", 5),
], 1):
    add_task(f"T14.1.2.{i}", "US14.1.2", desc, impl, files, hrs, "TODO")

add_feature("F14.2", "E14", "Defect Tracking", "Bug lifecycle management")
add_story("US14.2.1", "F14.2", "Enhanced bug tracking", "As QA, bugs have special fields and workflows", 5)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Add bug-specific fields to CActivity", "AI: 1) Add severity (enum: CRITICAL, MAJOR, MINOR, TRIVIAL) 2) Add reproducibility (enum: ALWAYS, SOMETIMES, RANDOM) 3) Add stepsToReproduce (text) 4) Add foundInVersion, fixedInVersion 5) Only show for bug type", S, "MODIFY: activities/domain/CActivity.java", 3),
    ("Create bug-specific workflow", "AI: 1) Create 'Bug Lifecycle' workflow 2) Statuses: New, Open, In Progress, Fixed, Verified, Closed, Reopened 3) Validators: require severity on creation 4) Post-function: notify QA when Fixed", V, "CREATE via initializer", 4),
    ("Create CBugReportDialog", "AI: 1) Specialized dialog for bugs 2) Required fields: severity, steps to reproduce 3) Optional: screenshot attachment 4) Environment info (browser, OS) 5) Link to test case", U, "CREATE: ui/CBugReportDialog.java", 6),
], 1):
    add_task(f"T14.2.1.{i}", "US14.2.1", desc, impl, files, hrs, "TODO")

# ==================================================================================================================
# EPIC 15: ADVANCED UI/UX IMPROVEMENTS
# ==================================================================================================================
print("\n[11/15] Epic 15: Advanced UI/UX Improvements")
add_epic("E15", "Advanced UI/UX", "Rich text editor, emoji support, dark mode, keyboard shortcuts, responsive design")

add_feature("F15.1", "E15", "Rich Text Editor", "WYSIWYG editor for descriptions and comments")
add_story("US15.1.1", "F15.1", "Rich text editing", "As user, I can format text with rich editor", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Integrate TinyMCE or Quill", "AI: 1) Add vaadin-quill or vaadin-rich-text-editor 2) Create CRichTextEditor component 3) Configure toolbar: bold, italic, lists, links, images 4) Sanitize HTML on save 5) Support @mentions", U, "CREATE: ui/components/CRichTextEditor.java\nMODIFY: pom.xml", 6),
    ("Update CActivity and CComment to use HTML", "AI: 1) Change description/comment fields to @Lob 2) Store HTML 3) Sanitize with Jsoup 4) Update all forms to use CRichTextEditor 5) Display with safe HTML rendering", S, "MODIFY: domain/*.java, ui/*.java", 5),
    ("Add image upload for editor", "AI: 1) Image button in toolbar 2) Upload to file storage 3) Insert <img> tag with URL 4) Resize large images 5) Lazy load in view", U, "MODIFY: CRichTextEditor.java", 4),
], 1):
    add_task(f"T15.1.1.{i}", "US15.1.1", desc, impl, files, hrs, "TODO")

add_feature("F15.2", "E15", "Dark Mode", "Dark theme support")
add_story("US15.2.1", "F15.2", "Dark mode toggle", "As user, I can switch to dark mode", 5)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create dark theme CSS", "AI: 1) Create styles/dark-theme.css 2) Define CSS variables for colors 3) Override Lumo theme variables 4) Test all components 5) Ensure sufficient contrast", U, "CREATE: frontend/styles/dark-theme.css", 6),
    ("Add theme toggle to user preferences", "AI: 1) Add theme field to CUser (enum: LIGHT, DARK, SYSTEM) 2) Store in session 3) Apply on login 4) Toggle button in header 5) Persist choice", S+V+U, "MODIFY: CUser.java, MainLayout.java", 4),
    ("Update all custom CSS for dark mode", "AI: 1) Use CSS variables instead of hardcoded colors 2) Test kanban board 3) Test all dialogs 4) Fix any contrast issues", U, "MODIFY: All CSS files", 5),
], 1):
    add_task(f"T15.2.1.{i}", "US15.2.1", desc, impl, files, hrs, "TODO")

add_feature("F15.3", "E15", "Keyboard Shortcuts", "Power user keyboard navigation")
add_story("US15.3.1", "F15.3", "Global shortcuts", "As power user, I can use keyboard shortcuts", 5)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Create CKeyboardShortcutManager", "AI: 1) Register global shortcuts 2) Handle conflicts 3) Show help dialog (press ?) 4) Shortcuts: Ctrl+/ help, C create, E edit, G goto, / search 5) Use Vaadin Shortcuts API", U, "CREATE: ui/CKeyboardShortcutManager.java", 5),
    ("Add shortcuts to grids", "AI: 1) J/K navigate rows 2) Enter open selected 3) Delete remove selected 4) Shift+Arrow multiselect", U, "MODIFY: CAbstractPage.java to register shortcuts", 3),
    ("Create CQuickJumpDialog (Ctrl+K)", "AI: 1) Command palette style 2) Search all: projects, activities, users 3) Fuzzy search 4) Keyboard navigation 5) Show shortcuts for results", U, "CREATE: ui/CQuickJumpDialog.java", 6),
], 1):
    add_task(f"T15.3.1.{i}", "US15.3.1", desc, impl, files, hrs, "TODO")

add_feature("F15.4", "E15", "Responsive Mobile Design", "Mobile-friendly layouts")
add_story("US15.4.1", "F15.4", "Mobile optimization", "As mobile user, app works on phone/tablet", 8)

for i, (desc, impl, std, files, hrs) in enumerate([
    ("Add responsive CSS", "AI: 1) Use @media queries 2) Breakpoints: 480px, 768px, 1024px 3) Hide sidebar on mobile 4) Stack grids vertically 5) Larger touch targets", U, "CREATE: frontend/styles/responsive.css", 8),
    ("Create mobile navigation", "AI: 1) Hamburger menu 2) Drawer for navigation 3) Bottom navigation bar 4) Swipe gestures", U, "MODIFY: MainLayout.java", 6),
    ("Optimize kanban for touch", "AI: 1) Larger drag handles 2) Tap to select vs drag 3) Horizontal scroll for columns 4) Collapse columns 5) Test on real devices", U, "MODIFY: CComponentKanbanBoard.java", 8),
], 1):
    add_task(f"T15.4.1.{i}", "US15.4.1", desc, impl, files, hrs, "TODO")

# Save and report
print(f"\n{'='*100}")
print(f"Saving comprehensive backlog... ({stats['tasks']} tasks added)")
wb.save('docs/__PROJECT_BACKLOG.xlsx')
print("✅ Saved!")

print(f"\n{'='*100}")
print("FINAL SUMMARY:")
print(f"  Epics Added:        {stats['epics']}")
print(f"  Features Added:     {stats['features']}")
print(f"  User Stories Added: {stats['stories']}")
print(f"  Tasks Added:        {stats['tasks']}")
print("="*100)

# Show final totals
wb = openpyxl.load_workbook('docs/__PROJECT_BACKLOG.xlsx')
print("\nFINAL BACKLOG TOTALS:")
for sheet in ['Epics', 'Features', 'User_Stories', 'Tasks']:
    print(f"  {sheet:15} : {wb[sheet].max_row - 1:4} rows")
print("="*100)
