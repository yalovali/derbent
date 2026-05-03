#!/usr/bin/env python3
"""
gen_screens_init.py
--------------------
Reads all *InitializerService.java files under the derbent src tree and
generates screens_init.xlsx with four sheets:
  Grid Entities | Detail Sections | Detail Lines | Page Entities

Each logical row is repeated once per project (6 projects total).
"""

import os
import re
import glob
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found – activate the venv first")
    raise

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SRC_ROOT = Path("/home/yasin/git/derbent/src/main/java")
OUT_FILE = Path("/home/yasin/git/derbent/src/main/resources/excel/screens_init.xlsx")

PROJECTS = [
    "Derbent PM Demo",
    "Derbent API Platform",
    "BAB Integration Program",
    "Mobile App Delivery",
    "Data & Analytics Platform",
    "Customer Portal Revamp",
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")

BASE_CAPTIONS = {
    "id": "#", "name": "Name", "description": "Description", "active": "Active",
    "createdDate": "Created Date", "lastModifiedDate": "Last Modified",
    "assignedTo": "Assigned To", "createdBy": "Created By", "project": "Project",
    "entityType": "Entity Type", "status": "Status", "color": "Color",
    "sortOrder": "Sort Order", "workflow": "Workflow", "company": "Company",
    "attributeNonDeletable": "Non Deletable", "priority": "Priority",
    "notes": "Notes", "results": "Results", "acceptanceCriteria": "Acceptance Criteria",
    "startDate": "Start Date", "dueDate": "Due Date", "completionDate": "Completion Date",
    "progressPercentage": "Progress %", "estimatedHours": "Estimated Hours",
    "actualHours": "Actual Hours", "remainingHours": "Remaining Hours",
    "storyPoints": "Story Points", "releaseVersion": "Release Version",
    "sprintOrder": "Sprint Order", "epicLink": "Epic Link",
    "parentFeature": "Parent Feature", "parentEpic": "Parent Epic",
    "businessValue": "Business Value", "riskLevel": "Risk Level",
    "complexity": "Complexity",
}

# Menu constants (from CInitializerServiceBase)
MENU_CONSTANTS = {
    "Menu_Order_CRM": "5",
    "Menu_Order_DEVELOPMENT": "9999.",
    "Menu_Order_FINANCE": "10",
    "Menu_Order_POLICIES": "60",
    "Menu_Order_PRODUCTS": "20",
    "Menu_Order_PROJECT": "1",
    "Menu_Order_ROLES": "400",
    "Menu_Order_SETUP": "400",
    "Menu_Order_SYSTEM": "500",
    "Menu_Order_TESTS": "15",
    "Menu_Order_TYPES": "130",
    "MenuTitle_CRM": "CRM",
    "MenuTitle_DEVELOPMENT": "Development.",
    "MenuTitle_FINANCE": "Finance",
    "MenuTitle_POLICIES": "Policies",
    "MenuTitle_PRODUCTS": "Products",
    "MenuTitle_PROJECT": "Project",
    "MenuTitle_ROLES": "Roles",
    "MenuTitle_SETUP": "Setup",
    "MenuTitle_STORAGE": "Storage",
    "MenuTitle_SYSTEM": "System",
    "MenuTitle_TESTS": "Tests",
    "MenuTitle_TYPES": "Types",
    "MenuTitle_TYPES_CRM": "CRM",
    "MenuTitle_TYPES_FINANCE": "Finance",
    "MenuTitle_TYPES_PRODUCTS": "Products",
    "MenuTitle_TYPES_PROJECT": "Project",
    "MenuTitle_TYPES_TESTS": "Tests",
}

# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------
def get_caption(field_name: str) -> str:
    if field_name in BASE_CAPTIONS:
        return BASE_CAPTIONS[field_name]
    words = re.sub(r'([A-Z])', r' \1', field_name).strip()
    return ' '.join(w.capitalize() for w in words.split())


def resolve_string_expr(expr: str, extra: dict = None) -> str:
    """Resolve a Java string concatenation expression to a Python string.
    Handles things like:  Menu_Order_PROJECT + ".2"
                          MenuTitle_PROJECT + ".Activities"
                          "literal"
    """
    # combine known constants
    constants = dict(MENU_CONSTANTS)
    if extra:
        constants.update(extra)

    def replace_const(m):
        token = m.group(0).strip('" ')
        return constants.get(token, token)

    # Remove outer quotes, handle concatenation
    parts = re.split(r'\s*\+\s*', expr.strip())
    result = []
    for part in parts:
        part = part.strip()
        if part.startswith('"') and part.endswith('"'):
            result.append(part[1:-1])
        elif part in constants:
            result.append(constants[part])
        else:
            # Try to keep as-is (might be unknown constant)
            result.append(part)
    return ''.join(result)


# ---------------------------------------------------------------------------
# Java file reading helpers
# ---------------------------------------------------------------------------
def read_file(path: str) -> str:
    with open(path, encoding="utf-8", errors="replace") as f:
        return f.read()


def extract_static_string(content: str, var_name: str) -> str | None:
    """Extract value of a private/protected static final String field."""
    pattern = rf'static\s+final\s+String\s+{re.escape(var_name)}\s*=\s*(.+?);'
    m = re.search(pattern, content, re.DOTALL)
    if not m:
        return None
    return resolve_string_expr(m.group(1).strip())


def extract_clazz(content: str) -> str | None:
    """Extract  static final Class<?> clazz = SomeClass.class;"""
    m = re.search(r'static\s+final\s+Class\s*<\??>\s+clazz\s*=\s*(\w+)\.class\s*;', content)
    if m:
        return m.group(1)
    # Some files use private static final
    m = re.search(r'Class\s*<\??>\s+clazz\s*=\s*(\w+)\.class\s*;', content)
    if m:
        return m.group(1)
    return None


def extract_column_fields(content: str, method_name: str = "setColumnFields") -> list[str]:
    """Extract List.of(...) argument from setColumnFields call."""
    pattern = rf'{re.escape(method_name)}\s*\(\s*List\.of\s*\(([^)]+)\)\s*\)'
    m = re.search(pattern, content, re.DOTALL)
    if not m:
        return []
    raw = m.group(1)
    # Extract quoted strings
    return re.findall(r'"([^"]+)"', raw)


def extract_screen_lines(content: str) -> list[dict]:
    """
    Parse the createBasicView method body and produce a list of screen-line dicts.
    Returns list of:
        {'type': 'section', 'section_name': '...'}
        {'type': 'field', 'field': '...', 'have_next': bool, 'width': '...'}
        {'type': 'special', 'description': '...'}
    """
    # Find createBasicView method body
    m = re.search(
        r'(public|private|protected)\s+static\s+CDetailSection\s+createBasicView\s*\([^)]*\)'
        r'\s*(?:throws\s+\w[\w,\s]*)?\s*\{',
        content
    )
    if not m:
        return []

    # Extract body by brace counting
    start = m.end()
    depth = 1
    pos = start
    while pos < len(content) and depth > 0:
        if content[pos] == '{':
            depth += 1
        elif content[pos] == '}':
            depth -= 1
        pos += 1
    body = content[start:pos - 1]

    lines = []
    current_section = "Description"  # default first section

    # Pattern: createSection("SectionName")
    # Pattern: createLineFromDefaults(clazz, "fieldName")
    # Pattern: createLineFromDefaults(clazz, "fieldName", true/false, "width")
    # Pattern: addDefaultSection  (attachments/comments/links/parentrelation)

    # We iterate through the body token by token using regex
    for call_match in re.finditer(
        r'(\w+(?:\.\w+)*)\s*\.\s*'
        r'(createSection|createLineFromDefaults|addDefaultSection|addDefaultChildrenSection)\s*'
        r'\(([^;]*?)\)\s*(?:;|,)',
        body,
        re.DOTALL
    ):
        caller = call_match.group(1)
        method = call_match.group(2)
        args_raw = call_match.group(3).strip()

        if method == 'createSection':
            # Extract section name
            sm = re.search(r'"([^"]*)"', args_raw)
            if sm:
                current_section = sm.group(1)
                lines.append({'type': 'section', 'section_name': current_section})

        elif method == 'createLineFromDefaults':
            # Args: clazz, "fieldName"  OR  clazz, "fieldName", true/false, "width"
            # Some calls: clazz, "fieldName", true, ""
            quoted = re.findall(r'"([^"]*)"', args_raw)
            bools = re.findall(r'\b(true|false)\b', args_raw)
            if not quoted:
                continue
            field_name = quoted[0]
            have_next = bools[0].lower() == 'true' if bools else False
            width = quoted[1] if len(quoted) > 1 else ''
            lines.append({
                'type': 'field',
                'field': field_name,
                'have_next': have_next,
                'width': width,
                'section': current_section,
            })

        elif method == 'addDefaultSection':
            # CAttachmentInitializerService.addDefaultSection -> "Attachments" section + "attachments" field
            # CCommentInitializerService.addDefaultSection    -> "Comments"    section + "comments"    field
            # CLinkInitializerService.addDefaultSection       -> "Links"       section + "links"       field
            # CParentRelationInitializerService.addDefaultSection -> "Parent Hierarchy" + "parentRelation"
            caller_lower = caller.lower()
            if 'attachment' in caller_lower:
                current_section = 'Attachments'
                lines.append({'type': 'section', 'section_name': 'Attachments'})
                lines.append({'type': 'field', 'field': 'attachments', 'have_next': False,
                               'width': '', 'section': 'Attachments'})
            elif 'comment' in caller_lower:
                current_section = 'Comments'
                lines.append({'type': 'section', 'section_name': 'Comments'})
                lines.append({'type': 'field', 'field': 'comments', 'have_next': False,
                               'width': '', 'section': 'Comments'})
            elif 'link' in caller_lower:
                current_section = 'Links'
                lines.append({'type': 'section', 'section_name': 'Links'})
                lines.append({'type': 'field', 'field': 'links', 'have_next': False,
                               'width': '', 'section': 'Links'})
            elif 'parentrelation' in caller_lower or 'parentrel' in caller_lower:
                current_section = 'Parent Hierarchy'
                lines.append({'type': 'section', 'section_name': 'Parent Hierarchy'})
                lines.append({'type': 'field', 'field': 'placeHolder_createComponentParent', 'have_next': False,
                               'width': '', 'section': 'Parent Hierarchy'})
            # else unknown, skip

        elif method == 'addDefaultChildrenSection':
            current_section = 'Children'
            lines.append({'type': 'section', 'section_name': 'Children'})
            lines.append({'type': 'field', 'field': 'placeHolder_createComponentParentChildren', 'have_next': False,
                           'width': '', 'section': 'Children'})

    # Also handle createScreenLines calls that inject base fields
    # CProjectItemInitializerService.createScreenLines -> Description section + base named fields + Type section with assignedTo/createdBy/project + entityType
    # CEntityOfProjectInitializerService.createScreenLines -> Description section + base named fields + Type section with assignedTo/createdBy/project
    # CEntityNamedInitializerService.createScreenLines -> Description section + base named fields
    lines = _inject_base_screen_lines(body, lines)

    return lines


def _inject_base_screen_lines(body: str, lines: list) -> list:
    """
    Inject base screen lines from createScreenLines calls at the beginning.
    These calls happen BEFORE the explicit lines in the method body.
    We prepend the corresponding fields.
    """
    # Look for createScreenLines calls
    # CProjectItemInitializerService.createScreenLines -> adds Description + assignedTo/createdBy/project + entityType
    # CEntityOfProjectInitializerService.createScreenLines -> adds Description + assignedTo/createdBy/project
    # CEntityNamedInitializerService.createScreenLines -> adds Description section + id,name,desc,active,createdDate,lastModifiedDate

    has_project_item = bool(re.search(r'CProjectItemInitializerService\s*\.\s*createScreenLines', body))
    has_entity_of_project = bool(re.search(r'CEntityOfProjectInitializerService\s*\.\s*createScreenLines', body))
    has_entity_named = bool(re.search(r'CEntityNamedInitializerService\s*\.\s*createScreenLines', body))

    # check if createTypeEntityView is called (different base path)
    has_type_entity_view = bool(re.search(r'createTypeEntityView\s*\(', body))

    base_lines = []

    if has_type_entity_view:
        # createTypeEntityView adds: Description section + id,name,desc,active,createdDate,lastModifiedDate
        # Then company, workflow, configSection with color,sortOrder,[extras],attributeNonDeletable,active
        # We let the parsed lines handle this since createTypeEntityView body is not in THIS file
        # So we inject standard named lines + company + workflow
        base_lines += _named_base_lines()
        base_lines.append({'type': 'field', 'field': 'company', 'have_next': False, 'width': '', 'section': 'Description'})
        base_lines.append({'type': 'field', 'field': 'workflow', 'have_next': False, 'width': '', 'section': 'Description'})
        # The rest (createSection, color, sortOrder, etc.) will be picked up by the explicit lines parser

    elif has_project_item:
        # adds Description + named fields + Type section + assignedTo, createdBy, project + entityType
        base_lines += _named_base_lines()
        base_lines.append({'type': 'section', 'section_name': 'Type'})
        base_lines.append({'type': 'field', 'field': 'assignedTo', 'have_next': False, 'width': '', 'section': 'Type'})
        base_lines.append({'type': 'field', 'field': 'createdBy', 'have_next': False, 'width': '', 'section': 'Type'})
        base_lines.append({'type': 'field', 'field': 'project', 'have_next': False, 'width': '', 'section': 'Type'})
        base_lines.append({'type': 'field', 'field': 'entityType', 'have_next': False, 'width': '', 'section': 'Type'})

    elif has_entity_of_project:
        # adds Description + named fields + Type section + assignedTo, createdBy, project
        base_lines += _named_base_lines()
        base_lines.append({'type': 'section', 'section_name': 'Type'})
        base_lines.append({'type': 'field', 'field': 'assignedTo', 'have_next': False, 'width': '', 'section': 'Type'})
        base_lines.append({'type': 'field', 'field': 'createdBy', 'have_next': False, 'width': '', 'section': 'Type'})
        base_lines.append({'type': 'field', 'field': 'project', 'have_next': False, 'width': '', 'section': 'Type'})

    elif has_entity_named:
        base_lines += _named_base_lines()

    return base_lines + lines


def _named_base_lines() -> list:
    """Return the standard Description section lines from CEntityNamedInitializerService."""
    return [
        {'type': 'section', 'section_name': 'Description'},
        {'type': 'field', 'field': 'id',               'have_next': True,  'width': '10%', 'section': 'Description'},
        {'type': 'field', 'field': 'name',             'have_next': False, 'width': '100%','section': 'Description'},
        {'type': 'field', 'field': 'description',      'have_next': False, 'width': '',    'section': 'Description'},
        {'type': 'field', 'field': 'active',           'have_next': False, 'width': '',    'section': 'Description'},
        {'type': 'field', 'field': 'createdDate',      'have_next': True,  'width': '',    'section': 'Description'},
        {'type': 'field', 'field': 'lastModifiedDate', 'have_next': True,  'width': '',    'section': 'Description'},
    ]


# ---------------------------------------------------------------------------
# Main parsing
# ---------------------------------------------------------------------------
def find_initializer_files() -> list[str]:
    pattern = str(SRC_ROOT / "**" / "*InitializerService.java")
    return sorted(glob.glob(pattern, recursive=True))


def should_skip(path: str) -> bool:
    """Skip abstract base classes and purely-data initializers that have no initialize() method."""
    skip_names = {
        'CInitializerServiceBase.java',
        'CEntityNamedInitializerService.java',
        'CEntityOfProjectInitializerService.java',
        'CProjectItemInitializerService.java',
        'CEntityOfCompanyInitializerService.java',
        'CEntityTypeInitializerService.java',
        'COneToOneRelationBaseInitializationService.java',
        'CAttachmentInitializerService.java',   # no standalone page (BAB skips)
        'CCommentInitializerService.java',       # no standalone page
        'CLinkInitializerService.java',          # no standalone page
        'CParentRelationInitializerService.java',# no standalone page
        'CEmailInitializerService.java',         # abstract
        # BAB files excluded for Derbent projects
        'CBabDeviceInitializerService.java',
        'CBabNodeInitializerService.java',
        'CBabPolicyActionInitializerService.java',
        'CBabPolicyActionMaskCANInitializerService.java',
        'CBabPolicyActionMaskFileInitializerService.java',
        'CBabPolicyActionMaskInitializerService.java',
        'CBabPolicyActionMaskROSInitializerService.java',
        'CBabPolicyFilterCANInitializerService.java',
        'CBabPolicyFilterCSVInitializerService.java',
        'CBabPolicyFilterInitializerService.java',
        'CBabPolicyFilterROSInitializerService.java',
        'CBabCanNodeInitializerService.java',
        'CBabFileInputNodeInitializerService.java',
        'CBabFileOutputNodeInitializerService.java',
        'CBabHttpServerNodeInitializerService.java',
        'CBabSyslogNodeInitializerService.java',
        'CBabTCPModbusNodeInitializerService.java',
        'CBabModbusNodeInitializerService.java',
        'CBabROSNodeInitializerService.java',
        'CBabPolicyRuleInitializerService.java',
        'CBabPolicybaseInitializerService.java',
        'CBabPolicyTriggerInitializerService.java',
        'CProject_BabInitializerService.java',
        'CSystemSettings_BabInitializerService.java',
        # Dashboard/extra files that don't follow standard pattern
        'CDashboardProjectTypeInitializerService.java',
        'CDashboardInterfaces_InitializerService.java',
        'CDashboardProject_BabInitializerService.java',
        'CScheduleTaskInitializerService.java',   # mostly system-level
        'CEntityInitializerService.java',
        'CEntityDBInitializerService.java',
        'CEmailQueuedInitializerService.java',    # may not have standard view
        'CEmailSentInitializerService.java',
        'CAgileEntityInitializerService.java',    # abstract agile base
        # Skip the main orchestrator
        'CScreensInitializerService.java',
    }
    return os.path.basename(path) in skip_names


def parse_initializer(path: str) -> dict | None:
    """Parse a single initializer file and return extracted info, or None to skip."""
    content = read_file(path)

    # Must have an initialize() static method
    if not re.search(r'public\s+static\s+void\s+initialize\s*\(', content):
        return None

    # Must have createGridEntity and createBasicView
    if not re.search(r'static\s+CGridEntity\s+createGridEntity', content) and \
       not re.search(r'private\s+static\s+CGridEntity\s+createGridEntity', content):
        return None

    clazz_name = extract_clazz(content)
    if not clazz_name:
        return None

    # Grid info
    col_fields = extract_column_fields(content, 'setColumnFields')
    editable_fields = extract_column_fields(content, 'setEditableColumnFields')

    # Check for attributeNone (None Grid)
    has_none_grid = bool(re.search(r'\.setAttributeNone\s*\(\s*true\s*\)', content))

    # Detail section info
    # Name comes from VIEW_NAME field (we use clazz_name as a fallback key)
    # The screen title / header text = VIEW_NAME which we can't get at parse time,
    # but initBase uses the clazz VIEW_NAME. We'll use clazz_name as a stand-in.
    screen_lines = extract_screen_lines(content)

    # Page entity info
    menu_title = extract_static_string(content, 'menuTitle')
    menu_order = extract_static_string(content, 'menuOrder')
    page_title = extract_static_string(content, 'pageTitle')
    page_description = extract_static_string(content, 'pageDescription')

    # Some files override menuTitle/menuOrder inside initialize() rather than as a field
    # e.g. CSystemSettings_DerbentInitializerService, CGridEntityInitializerService
    if menu_title is None or menu_order is None:
        init_body_m = re.search(
            r'public\s+static\s+void\s+initialize\s*\([^)]*\)\s*(?:throws[^{]*)?\{',
            content
        )
        if init_body_m:
            init_body = content[init_body_m.end(): init_body_m.end() + 3000]
            if menu_title is None:
                mt_m = re.search(r'"([^"]+)"\s*,\s*pageTitle', init_body)
                if mt_m is None:
                    # look for local variable menuTitle assignment
                    mt_m2 = re.search(r'String\s+menuTitle\s*=\s*"([^"]+)"', init_body)
                    if mt_m2:
                        menu_title = mt_m2.group(1)
                else:
                    menu_title = mt_m.group(1)
            if menu_order is None:
                mo_m = re.search(r'String\s+menuOrder\s*=\s*"([^"]+)"', init_body)
                if mo_m:
                    menu_order = mo_m.group(1)

    # Try to extract menuTitle from initBase call directly
    # initBase(clazz, project, ..., menuTitle, pageTitle, pageDescription, ...)
    # We look for the literal strings in the initBase call
    if menu_title is None or menu_order is None or page_title is None:
        initbase_m = re.search(r'initBase\s*\((.+?)\)\s*;', content, re.DOTALL)
        if initbase_m:
            ib_args = initbase_m.group(1)
            # Split by comma respecting nesting
            ib_parts = _split_args(ib_args)
            # initBase(clazz, project, gridEntityService, detailSectionService, pageEntityService,
            #          detailSection, grid, menuTitle, pageTitle, pageDescription, showInQuickToolbar, order, customizer)
            # indices:   0        1        2                 3                     4
            #            5              6     7           8           9            10                  11    12
            if len(ib_parts) >= 12:
                if menu_title is None:
                    menu_title = _try_resolve_expr(ib_parts[7].strip(), content)
                if page_title is None:
                    page_title = _try_resolve_expr(ib_parts[8].strip(), content)
                if menu_order is None:
                    menu_order = _try_resolve_expr(ib_parts[11].strip(), content)

    # Derive page service name: CPageService + clazz_name[1:]
    # e.g. CActivity -> CPageServiceActivity
    bare = clazz_name[1:] if clazz_name.startswith('C') else clazz_name
    page_service = f"CPageService{bare}"

    # Build detail section name: same as VIEW_NAME pattern which is clazz_name without 'C'
    # The actual VIEW_NAME is defined in the domain class, not the initializer.
    # We use clazz_name as a key and note the name will be the VIEW_NAME at runtime.
    # For the Excel we use clazz_name as the section name identifier.
    detail_section_name = clazz_name

    return {
        'clazz': clazz_name,
        'detail_section_name': detail_section_name,
        'col_fields': col_fields,
        'editable_fields': editable_fields,
        'has_none_grid': has_none_grid,
        'screen_lines': screen_lines,
        'menu_title': menu_title or '',
        'menu_order': menu_order or '',
        'page_title': page_title or clazz_name,
        'page_description': page_description or '',
        'page_service': page_service,
        # service bean name for grid: pattern is clazz + "Service"
        'data_service_bean': clazz_name + 'Service',
        'file': path,
    }


def _split_args(s: str) -> list[str]:
    """Split a comma-separated argument list respecting parentheses depth."""
    parts = []
    depth = 0
    current = []
    for ch in s:
        if ch == '(':
            depth += 1
            current.append(ch)
        elif ch == ')':
            depth -= 1
            current.append(ch)
        elif ch == ',' and depth == 0:
            parts.append(''.join(current))
            current = []
        else:
            current.append(ch)
    if current:
        parts.append(''.join(current))
    return parts


def _try_resolve_expr(expr: str, content: str) -> str | None:
    """Try to resolve a Java expression to a string, consulting file-level constants."""
    expr = expr.strip()
    if expr.startswith('"') and expr.endswith('"'):
        return expr[1:-1]
    # Could be a constant reference like MenuTitle_PROJECT + ".Activities"
    # Build a local constants dict including file-level strings
    local_consts = dict(MENU_CONSTANTS)
    # Extract all string constants from the file
    for m in re.finditer(
        r'static\s+final\s+String\s+(\w+)\s*=\s*([^;]+);',
        content, re.DOTALL
    ):
        var_name = m.group(1)
        val_expr = m.group(2).strip()
        # Only resolve simple ones
        if '"' in val_expr or any(k in val_expr for k in MENU_CONSTANTS):
            resolved = resolve_string_expr(val_expr, local_consts)
            local_consts[var_name] = resolved

    return resolve_string_expr(expr, local_consts)


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------
def apply_header_style(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def auto_width(ws, min_width=10, max_width=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def bool_str(v: bool) -> str:
    return "TRUE" if v else "FALSE"


def generate_excel(parsed: list[dict]) -> None:
    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Grid Entities
    # -----------------------------------------------------------------------
    ws_grid = wb.active
    ws_grid.title = "Grid Entities"
    grid_headers = [
        "project", "name", "Data Service Bean", "Column Fields",
        "Editable Column Fields", "None Grid"
    ]
    ws_grid.append(grid_headers)
    for cell in ws_grid[1]:
        apply_header_style(cell)

    grid_rows = []
    for item in parsed:
        col_str = ",".join(item['col_fields'])
        editable_str = ",".join(item['editable_fields']) if item['editable_fields'] else ""
        none_grid = bool_str(item['has_none_grid'])
        for proj in PROJECTS:
            grid_rows.append([
                proj, item['clazz'], item['data_service_bean'],
                col_str, editable_str, none_grid
            ])
    for row in grid_rows:
        ws_grid.append(row)
    auto_width(ws_grid)

    # -----------------------------------------------------------------------
    # Sheet 2: Detail Sections
    # -----------------------------------------------------------------------
    ws_ds = wb.create_sheet("Detail Sections")
    ds_headers = [
        "project", "name", "Entity Type", "Screen Title",
        "Header Text", "Default Section", "Non Deletable"
    ]
    ws_ds.append(ds_headers)
    for cell in ws_ds[1]:
        apply_header_style(cell)

    ds_rows = []
    for item in parsed:
        for proj in PROJECTS:
            ds_rows.append([
                proj,
                item['detail_section_name'],
                item['clazz'],          # Entity Type = class name
                item['clazz'],          # Screen Title (VIEW_NAME = class name)
                item['clazz'],          # Header Text
                "",                     # Default Section (null)
                "TRUE",                 # Non Deletable
            ])
    for row in ds_rows:
        ws_ds.append(row)
    auto_width(ws_ds)

    # -----------------------------------------------------------------------
    # Sheet 3: Detail Lines
    # -----------------------------------------------------------------------
    ws_dl = wb.create_sheet("Detail Lines")
    dl_headers = [
        "project", "Detail Section", "Entity Property", "Field Caption",
        "Relation Field Name", "Item Order", "Is Hidden", "Is Readonly",
        "Is Required", "Have Next One On Same Line", "Section As Tab",
        "Width", "Section Name", "Data Provider Bean", "Default Value",
        "Field Description", "Max Length", "Related Entity Type"
    ]
    ws_dl.append(dl_headers)
    for cell in ws_dl[1]:
        apply_header_style(cell)

    dl_rows = []
    for item in parsed:
        screen_lines = item['screen_lines']
        order = 0
        current_section = ""
        for sl in screen_lines:
            if sl['type'] == 'section':
                # Section start row
                current_section = sl['section_name']
                for proj in PROJECTS:
                    dl_rows.append([
                        proj,
                        item['detail_section_name'],  # Detail Section
                        "Section Start",               # Entity Property
                        sl['section_name'],            # Field Caption = section name for section rows
                        "Section Start",               # Relation Field Name
                        order,                         # Item Order
                        "FALSE",                       # Is Hidden
                        "FALSE",                       # Is Readonly
                        "FALSE",                       # Is Required
                        "FALSE",                       # Have Next One On Same Line
                        "FALSE",                       # Section As Tab
                        "",                            # Width
                        sl['section_name'],            # Section Name
                        "",                            # Data Provider Bean
                        "",                            # Default Value
                        "",                            # Field Description
                        "",                            # Max Length
                        "",                            # Related Entity Type
                    ])
                order += 1
            elif sl['type'] == 'field':
                field_name = sl['field']
                caption = get_caption(field_name)
                have_next = bool_str(sl.get('have_next', False))
                width = sl.get('width', '')
                for proj in PROJECTS:
                    dl_rows.append([
                        proj,
                        item['detail_section_name'],  # Detail Section
                        field_name,                    # Entity Property
                        caption,                       # Field Caption
                        "This Class",                  # Relation Field Name
                        order,                         # Item Order
                        "FALSE",                       # Is Hidden
                        "FALSE",                       # Is Readonly
                        "FALSE",                       # Is Required
                        have_next,                     # Have Next One On Same Line
                        "FALSE",                       # Section As Tab
                        width,                         # Width
                        current_section,               # Section Name
                        "",                            # Data Provider Bean
                        "",                            # Default Value
                        "",                            # Field Description
                        "",                            # Max Length
                        "",                            # Related Entity Type
                    ])
                order += 1

    for row in dl_rows:
        ws_dl.append(row)
    auto_width(ws_dl)

    # -----------------------------------------------------------------------
    # Sheet 4: Page Entities
    # -----------------------------------------------------------------------
    ws_pe = wb.create_sheet("Page Entities")
    pe_headers = [
        "project", "name", "Menu Title", "Menu Order", "Page Title",
        "Page Service", "Icon", "Grid Entity", "Content"
    ]
    ws_pe.append(pe_headers)
    for cell in ws_pe[1]:
        apply_header_style(cell)

    pe_rows = []
    for item in parsed:
        for proj in PROJECTS:
            pe_rows.append([
                proj,
                item['clazz'],                      # name
                item['menu_title'] or "",
                item['menu_order'] or "",
                item['page_title'] or item['clazz'],
                item['page_service'],
                "",                                  # Icon (determined at runtime)
                item['clazz'],                       # Grid Entity name
                "",                                  # Content
            ])
    for row in pe_rows:
        ws_pe.append(row)
    auto_width(ws_pe)

    # -----------------------------------------------------------------------
    # Freeze header row on all sheets
    # -----------------------------------------------------------------------
    for ws in [ws_grid, ws_ds, ws_dl, ws_pe]:
        ws.freeze_panes = "A2"

    wb.save(OUT_FILE)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    files = find_initializer_files()
    print(f"Found {len(files)} initializer files")

    parsed = []
    skipped = []
    errors = []

    for path in files:
        if should_skip(path):
            skipped.append(os.path.basename(path))
            continue
        try:
            result = parse_initializer(path)
            if result is None:
                skipped.append(os.path.basename(path))
            else:
                parsed.append(result)
        except Exception as e:
            errors.append((os.path.basename(path), str(e)))

    print(f"Parsed: {len(parsed)} | Skipped: {len(skipped)} | Errors: {len(errors)}")
    if errors:
        for fname, err in errors:
            print(f"  ERROR {fname}: {err}")

    # De-duplicate by clazz name (some might be parsed twice)
    seen = set()
    unique = []
    for item in parsed:
        if item['clazz'] not in seen:
            seen.add(item['clazz'])
            unique.append(item)
    print(f"Unique entities: {len(unique)}")

    generate_excel(unique)

    # Summary
    n = len(unique)
    n_proj = len(PROJECTS)
    print(f"\nExcel written to: {OUT_FILE}")
    print(f"  Grid Entities  rows: {n * n_proj}")
    print(f"  Detail Sections rows: {n * n_proj}")

    # Count detail lines
    total_dl = 0
    for item in unique:
        total_dl += len(item['screen_lines'])
    print(f"  Detail Lines   rows: {total_dl * n_proj}  (from {total_dl} unique lines across {n} entities)")
    print(f"  Page Entities  rows: {n * n_proj}")

    size = OUT_FILE.stat().st_size
    print(f"\nFile size: {size:,} bytes ({size/1024:.1f} KB)")


if __name__ == "__main__":
    main()
