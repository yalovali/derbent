#!/usr/bin/env python3
"""
Generate Code Quality Matrix for Derbent Project
This script creates a comprehensive Excel spreadsheet that tracks code quality
patterns, rules, and guidelines compliance across all classes in the codebase.
"""

import os
import re
import subprocess
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Base directory for the project
BASE_DIR = Path(__file__).parent.parent.parent.resolve()
SRC_DIR = BASE_DIR / "src/main/java/tech/derbent"

# Quality dimensions to check
QUALITY_DIMENSIONS = [
    # Naming and Structure
    ("C-Prefix Naming", "Class name follows C-prefix convention"),
    ("Package Structure", "Correct package structure (domain/service/view)"),
    
    # Entity Patterns
    ("Entity Annotations", "Has @Entity, @Table, @AttributeOverride (for entities)"),
    ("Entity Constants", "Has all 5 required constants (DEFAULT_COLOR, DEFAULT_ICON, etc.)"),
    ("Extends Base Class", "Extends appropriate base class (CEntityDB, CProjectItem, etc.)"),
    ("Interface Implementation", "Implements required interfaces correctly"),
    
    # Copy Pattern (NEW - 2026-01-17)
    ("copyEntityTo() Override", "Overrides copyEntityTo() if has entity-specific fields"),
    ("Calls super.copyEntityTo()", "Calls super.copyEntityTo(target, options) FIRST"),
    ("No Manual Interface Calls", "Does NOT manually call IHasComments/IHasAttachments copy methods"),
    ("Interface Copy Method", "Interface has static copy*To() method (if new interface)"),
    
    # Field Annotations and Validation
    ("@AMetaData Annotations", "All fields have @AMetaData with proper attributes"),
    ("Validation Annotations", "Has @NotNull, @NotBlank, @Size where appropriate"),
    ("Column Annotations", "Proper @Column, @JoinColumn annotations"),
    ("Fetch Strategy", "Uses LAZY fetch for collections/relationships"),
    
    # Constructor and Initialization
    ("Default Constructor", "Has no-arg constructor for JPA"),
    ("Named Constructor", "Has constructor(name, project) for creation"),
    ("initializeDefaults()", "Implements initializeDefaults() method"),
    
    # Repository Patterns
    ("Repository Interface", "Has repository interface extending correct base"),
    ("findById Override", "Overrides findById with JOIN FETCH for lazy fields"),
    ("Query Patterns", "Uses triple-quoted queries with #{#entityName}"),
    ("ORDER BY Clause", "All list queries have ORDER BY"),
    
    # Service Patterns
    ("Service Annotations", "Has @Service, @PreAuthorize, @PermitAll"),
    ("Service Base Class", "Extends appropriate service base class"),
    ("Stateless Service", "No instance state (multi-user safe)"),
    ("getEntityClass()", "Implements getEntityClass() method"),
    ("getInitializerService()", "Implements getInitializerService() method"),
    
    # Initializer Patterns
    ("Initializer Structure", "Has InitializerService extending CInitializerServiceBase"),
    ("createBasicView()", "Implements createBasicView() method"),
    ("createGridEntity()", "Implements createGridEntity() method"),
    ("initialize()", "Implements initialize() method"),
    ("initializeSample()", "Implements initializeSample() method"),
    ("CDataInitializer Registration", "Registered in CDataInitializer"),
    
    # Page Service Patterns
    ("Page Service", "Has CPageService if needed (workflow/sprint)"),
    ("Page Service Interfaces", "Implements correct page service interfaces"),
    
    # Exception Handling
    ("Exception Pattern", "Uses Check.notNull, proper exception handling"),
    ("User Exception Handling", "UI handlers use CNotificationService.showException"),
    
    # Logging
    ("Logger Field", "Has static final Logger with correct class"),
    ("Logging Pattern", "Follows ANSI logging format standards"),
    ("Log Levels", "Appropriate log levels (DEBUG, INFO, WARN, ERROR)"),
    
    # Interface Implementations
    ("IHasAttachments", "Proper getAttachments/setAttachments if applicable"),
    ("IHasComments", "Proper getComments/setComments if applicable"),
    ("IHasStatusAndWorkflow", "Proper workflow methods if applicable"),
    
    # Code Quality
    ("Getter/Setter Pattern", "Setters call updateLastModified()"),
    ("No Raw Types", "All generic types properly parameterized"),
    ("Constants Naming", "Constants are static final SCREAMING_SNAKE_CASE"),
    
    # Testing
    ("Unit Tests", "Has corresponding *Test class"),
    ("Integration Tests", "Has service/repository tests"),
    ("UI Tests", "Has Playwright tests if has view"),
    
    # Documentation
    ("JavaDoc", "Has class-level JavaDoc"),
    ("Method Documentation", "Complex methods documented"),
    ("Implementation Doc", "Has implementation markdown if complex"),
    
    # Security
    ("Access Control", "Proper @PreAuthorize/@RolesAllowed annotations"),
    ("Tenant Context", "Uses session service for company/project context"),
    
    # Formatting
    ("Code Formatting", "Follows eclipse-formatter.xml (4 spaces)"),
    ("Import Organization", "Clean imports, no wildcards"),
]

def get_class_info(class_path):
    """Extract information about a class."""
    parts = class_path.split('.')
    module = parts[2] if len(parts) > 2 else ""
    layer = parts[-2] if len(parts) > 1 else ""  # domain, service, view
    class_name = parts[-1]
    
    # Determine file path - reconstruct from class path
    # tech.derbent.api.companies.domain.CCompany -> src/main/java/tech/derbent/api/companies/domain/CCompany.java
    rel_path = class_path.replace('.', '/') + ".java"
    file_path = BASE_DIR / "src/main/java" / rel_path
    
    return {
        'full_path': class_path,
        'module': module,
        'layer': layer,
        'class_name': class_name,
        'file_path': file_path if file_path.exists() else None,
    }

def count_field_violations(content):
    """Count fields missing @AMetaData or validation annotations."""
    violations = []
    
    # Find all field declarations (not methods - methods have '(' after the name)
    # Pattern matches: @Annotation private Type fieldName; or = value;
    # But NOT: public Type methodName(...) { }
    field_pattern = r'@\w+.*?\n\s*(?:private|protected|public)\s+(?:static\s+)?(?:final\s+)?(\w+(?:<[^>]+>)?)\s+(\w+)\s*[=;]'
    fields = re.finditer(field_pattern, content, re.MULTILINE)
    
    for match in fields:
        field_type = match.group(1)
        field_name = match.group(2)
        
        # Get the annotations before this field
        start = max(0, match.start() - 500)
        field_section = content[start:match.end()]
        
        # Calculate line number
        line_number = content[:match.start()].count('\n') + 1
        
        # Check for @AMetaData
        if '@AMetaData' not in field_section and 'static' not in field_section and 'LOGGER' not in field_name:
            violations.append(('missing_ametadata', field_name, line_number, field_type))
        
        # Check for validation on non-collection fields
        if 'Set<' not in field_type and 'List<' not in field_type and 'Map<' not in field_type:
            if '@NotNull' not in field_section and '@NotBlank' not in field_section and '@Size' not in field_section:
                if 'String' in field_type or 'Integer' in field_type or 'Long' in field_type:
                    violations.append(('missing_validation', field_name, line_number, field_type))
    
    return violations

def count_missing_constants(content, required_constants):
    """Count missing required constants."""
    violations = []
    for const in required_constants:
        if const not in content:
            violations.append(('missing_constant', const))
    return violations

def analyze_class_file(file_path):
    """Analyze a Java file for quality patterns."""
    if not file_path or not file_path.exists():
        return {}
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return {}
    
    analysis = {}
    analysis['content'] = content  # Store content for detailed violation counting
    analysis['file_path'] = file_path
    
    # Check naming conventions
    analysis['c_prefix'] = re.search(r'(class|interface)\s+[CI]\w+', content) is not None
    
    # Check entity annotations
    analysis['entity_annotation'] = '@Entity' in content
    analysis['table_annotation'] = '@Table' in content
    analysis['attribute_override'] = '@AttributeOverride' in content
    
    # Check constants
    analysis['default_color'] = 'DEFAULT_COLOR' in content
    analysis['default_icon'] = 'DEFAULT_ICON' in content
    analysis['entity_title_singular'] = 'ENTITY_TITLE_SINGULAR' in content
    analysis['entity_title_plural'] = 'ENTITY_TITLE_PLURAL' in content
    analysis['view_name'] = 'VIEW_NAME' in content
    
    # Count field violations
    analysis['field_violations'] = count_field_violations(content)
    
    # Check annotations
    analysis['ametadata'] = '@AMetaData' in content
    analysis['notnull'] = '@NotNull' in content or '@NotBlank' in content
    analysis['service_annotation'] = '@Service' in content
    analysis['preauthorize'] = '@PreAuthorize' in content
    
    # Check methods
    analysis['initialize_defaults'] = 'initializeDefaults()' in content
    analysis['get_entity_class'] = 'getEntityClass()' in content
    analysis['create_basic_view'] = 'createBasicView' in content
    analysis['create_grid_entity'] = 'createGridEntity' in content
    analysis['initialize_sample'] = 'initializeSample' in content
    
    # Check copy pattern (NEW - 2026-01-17)
    analysis['copyentity_override'] = re.search(r'protected\s+void\s+copyEntityTo\s*\(', content) is not None
    analysis['super_copyentity'] = 'super.copyEntityTo(target, options)' in content or 'super.copyEntityTo(target,options)' in content
    analysis['no_manual_ihascomments'] = 'IHasComments.copyCommentsTo(' not in content or 'IHasComments.copyCommentsTo(this, target' not in content
    analysis['no_manual_ihasattachments'] = 'IHasAttachments.copyAttachmentsTo(' not in content or 'IHasAttachments.copyAttachmentsTo(this, target' not in content
    analysis['interface_copy_method'] = re.search(r'static\s+boolean\s+copy\w+To\s*\(.*CEntityDB', content) is not None
    
    # Check logging
    analysis['logger'] = re.search(r'Logger\s+LOGGER', content) is not None
    
    # Check base class extension
    analysis['extends'] = re.search(r'extends\s+\w+', content) is not None
    
    # Check interfaces
    analysis['implements'] = 'implements' in content
    
    # Check JOIN FETCH
    analysis['join_fetch'] = 'JOIN FETCH' in content
    
    # Check ORDER BY
    analysis['order_by'] = 'ORDER BY' in content
    
    # Check JavaDoc
    analysis['javadoc'] = '/**' in content
    
    # Check tests exist
    test_file = file_path.parent / f"{file_path.stem}Test.java"
    analysis['has_test'] = test_file.exists()
    
    return analysis

def count_violations(class_info, analysis, dimension_key):
    """Count violations for a quality dimension. Returns 'N/A' or a number (0-10+)."""
    class_name = class_info['class_name']
    layer = class_info['layer']
    content = analysis.get('content', '')
    
    # Special handling for different class types
    is_entity = layer == 'domain'
    is_service = layer == 'service' and 'Service' in class_name
    is_repository = 'Repository' in class_name
    is_initializer = 'Initializer' in class_name
    is_page_service = 'PageService' in class_name
    is_view = layer == 'view'
    is_exception = 'Exception' in class_name
    is_config = 'config' in class_info['module']
    
    # Map dimension to checks - return violation count or "N/A"
    if dimension_key == "C-Prefix Naming":
        if analysis.get('c_prefix'):
            return 0
        return 1  # Class name doesn't follow C-prefix
    
    elif dimension_key == "Entity Annotations":
        if not is_entity:
            return "N/A"
        violations = 0
        if not analysis.get('entity_annotation'):
            violations += 1
        if not analysis.get('table_annotation'):
            violations += 1
        if not analysis.get('attribute_override'):
            violations += 1
        return violations
    
    elif dimension_key == "Entity Constants":
        if not is_entity:
            return "N/A"
        violations = 0
        if not analysis.get('default_color'):
            violations += 1
        if not analysis.get('default_icon'):
            violations += 1
        if not analysis.get('entity_title_singular'):
            violations += 1
        if not analysis.get('entity_title_plural'):
            violations += 1
        if not analysis.get('view_name'):
            violations += 1
        return violations
    
    elif dimension_key == "@AMetaData Annotations":
        if not is_entity:
            return "N/A"
        # Count fields missing @AMetaData
        field_violations = analysis.get('field_violations', [])
        missing_ametadata = [v for v in field_violations if v[0] == 'missing_ametadata']
        return min(len(missing_ametadata), 10)
    
    elif dimension_key == "Validation Annotations":
        if not is_entity:
            return "N/A"
        # Count fields missing validation
        field_violations = analysis.get('field_violations', [])
        missing_validation = [v for v in field_violations if v[0] == 'missing_validation']
        return min(len(missing_validation), 10)
    
    elif dimension_key == "Service Annotations":
        if not is_service:
            return "N/A"
        violations = 0
        if not analysis.get('service_annotation'):
            violations += 1
        if not analysis.get('preauthorize'):
            violations += 1
        return violations
    
    elif dimension_key == "Repository Interface":
        if not is_repository:
            return "N/A"
        return 0 if class_info.get('file_path') else 1
    
    elif dimension_key == "findById Override":
        if not is_repository:
            return "N/A"
        return 0 if analysis.get('join_fetch') else 1
    
    elif dimension_key == "ORDER BY Clause":
        if not is_repository:
            return "N/A"
        # Count queries without ORDER BY
        query_pattern = r'@Query\s*\(\s*"""(.*?)"""'
        queries = re.findall(query_pattern, content, re.DOTALL)
        violations = 0
        for query in queries:
            if 'SELECT' in query.upper() and 'ORDER BY' not in query.upper():
                violations += 1
                if violations >= 10:
                    break
        return violations
    
    elif dimension_key == "Logger Field":
        if is_config or is_exception:
            return "N/A"
        return 0 if analysis.get('logger') else 1
    
    elif dimension_key == "Initializer Structure":
        if not is_initializer:
            return "N/A"
        return 0 if class_info.get('file_path') else 1
    
    elif dimension_key == "createBasicView()":
        if not is_initializer:
            return "N/A"
        return 0 if analysis.get('create_basic_view') else 1
    
    elif dimension_key == "createGridEntity()":
        if not is_initializer:
            return "N/A"
        return 0 if analysis.get('create_grid_entity') else 1
    
    elif dimension_key == "initializeSample()":
        if not is_initializer:
            return "N/A"
        return 0 if analysis.get('initialize_sample') else 1
    
    elif dimension_key == "Unit Tests":
        if is_config or is_exception:
            return "N/A"
        return 0 if analysis.get('has_test') else 1
    
    elif dimension_key == "JavaDoc":
        # Count public methods without JavaDoc
        method_pattern = r'\n\s*public\s+(?!class|interface)\w+[^{]*\{'
        public_methods = re.findall(method_pattern, content)
        violations = 0
        for _ in public_methods:
            # Simple heuristic: if file has some JavaDoc, assume some compliance
            pass
        return 0 if analysis.get('javadoc') else 1
    
    elif dimension_key == "Extends Base Class":
        if is_exception or is_config:
            return "N/A"
        return 0 if analysis.get('extends') else 1
    
    elif dimension_key == "Interface Implementation":
        if not (is_entity or is_service or is_page_service):
            return "N/A"
        # Can't easily count violations here, return binary
        return 0 if analysis.get('implements') else 1
    
    # Copy Pattern Checks (NEW - 2026-01-17)
    elif dimension_key == "copyEntityTo() Override":
        if not is_entity:
            return "N/A"
        # This is optional, so return 0 if present or N/A otherwise
        return 0 if analysis.get('copyentity_override') else "N/A"
    
    elif dimension_key == "Calls super.copyEntityTo()":
        if not is_entity or not analysis.get('copyentity_override'):
            return "N/A"
        return 0 if analysis.get('super_copyentity') else 1
    
    elif dimension_key == "No Manual Interface Calls":
        if not is_entity or not analysis.get('copyentity_override'):
            return "N/A"
        violations = 0
        no_comments = analysis.get('no_manual_ihascomments', True)
        no_attachments = analysis.get('no_manual_ihasattachments', True)
        if not no_comments:
            violations += 1
        if not no_attachments:
            violations += 1
        return violations
    
    elif dimension_key == "Interface Copy Method":
        # Only check for actual interface files
        if 'interface' not in class_name.lower() or not class_name.startswith('I'):
            return "N/A"
        return 0 if analysis.get('interface_copy_method') else 1
    
    elif dimension_key == "initializeDefaults()":
        if not is_entity:
            return "N/A"
        return 0 if analysis.get('initialize_defaults') else 1
    
    elif dimension_key == "getEntityClass()":
        if not is_service:
            return "N/A"
        return 0 if analysis.get('get_entity_class') else 1
    
    else:
        # For dimensions we can't automatically detect, return N/A
        return "N/A"

def create_excel_matrix(classes_file, output_file):
    """Create comprehensive Excel quality matrix."""
    
    # Read all classes
    with open(classes_file, 'r') as f:
        class_paths = [line.strip() for line in f if line.strip()]
    
    print(f"Found {len(class_paths)} classes to analyze")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Code Quality Matrix"
    
    # Define colors - gradient from green (0) to yellow (1-3) to red (4+)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    zero_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green - perfect
    low_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow - 1-3 violations
    medium_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid") # Orange - 4-6 violations
    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Red - 7+ violations
    na_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")     # Gray - N/A
    
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header row 1: Main categories
    ws.merge_cells('A1:E1')
    ws['A1'] = "Class Information"
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    col_offset = 6  # Start quality dimensions at column F
    
    # Header row 2: Column names
    headers = ["Class Name", "Module", "Layer", "File Path", "Category"]
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add quality dimension headers
    for idx, (dim_name, dim_desc) in enumerate(QUALITY_DIMENSIONS, col_offset):
        cell = ws.cell(row=1, column=idx)
        cell.value = dim_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=90)
        cell.border = border
        
        # Add description in row 2
        cell2 = ws.cell(row=2, column=idx)
        cell2.value = dim_desc
        cell2.fill = header_fill
        cell2.font = Font(size=8, color="FFFFFF")
        cell2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell2.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 20
    
    for idx in range(col_offset, col_offset + len(QUALITY_DIMENSIONS)):
        ws.column_dimensions[get_column_letter(idx)].width = 4
    
    # Freeze panes
    ws.freeze_panes = 'F3'
    
    # Track detailed violations for 10+ violations sheet
    detailed_violations = []
    
    # Process each class
    row = 3
    for class_path in class_paths:
        if row % 50 == 0:
            print(f"Processing row {row}/{len(class_paths) + 2}...")
        
        class_info = get_class_info(class_path)
        analysis = analyze_class_file(class_info['file_path'])
        
        # Determine category
        category = "Other"
        if class_info['layer'] == 'domain':
            category = "Entity"
        elif 'Service' in class_info['class_name']:
            if 'Initializer' in class_info['class_name']:
                category = "Initializer"
            elif 'PageService' in class_info['class_name']:
                category = "Page Service"
            else:
                category = "Service"
        elif 'Repository' in class_info['class_name']:
            category = "Repository"
        elif class_info['layer'] == 'view':
            category = "View"
        elif 'Exception' in class_info['class_name']:
            category = "Exception"
        elif 'config' in class_info['module']:
            category = "Configuration"
        
        # Write class info
        ws.cell(row=row, column=1, value=class_info['class_name']).border = border
        ws.cell(row=row, column=2, value=class_info['module']).border = border
        ws.cell(row=row, column=3, value=class_info['layer']).border = border
        ws.cell(row=row, column=4, value=str(class_info['file_path']) if class_info['file_path'] else 'N/A').border = border
        ws.cell(row=row, column=5, value=category).border = border
        
        # Write quality dimension violation counts and collect detailed violations
        for idx, (dim_name, _) in enumerate(QUALITY_DIMENSIONS, col_offset):
            violation_count = count_violations(class_info, analysis, dim_name)
            cell = ws.cell(row=row, column=idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            if violation_count == "N/A":
                cell.value = "-"
                cell.fill = na_fill
            else:
                # Display count, cap at 10
                display_value = str(violation_count) if violation_count < 10 else "10+"
                cell.value = display_value
                
                # Color code based on violation count
                if violation_count == 0:
                    cell.fill = zero_fill
                    cell.font = Font(color="006100", bold=True)  # Dark green
                elif violation_count <= 3:
                    cell.fill = low_fill
                    cell.font = Font(color="9C5700")  # Dark yellow
                elif violation_count <= 6:
                    cell.fill = medium_fill
                    cell.font = Font(color="9C2700")  # Dark orange
                else:
                    cell.fill = high_fill
                    cell.font = Font(color="9C0006", bold=True)  # Dark red
                
                # Collect detailed violations for classes with 10+ violations
                if violation_count >= 10:
                    # Get detailed field violations
                    field_violations = analysis.get('field_violations', [])
                    if dim_name == "@AMetaData Annotations":
                        missing = [v for v in field_violations if v[0] == 'missing_ametadata']
                        for violation_type, field_name, line_num, field_type in missing:
                            detailed_violations.append({
                                'class': class_info['class_name'],
                                'file': str(class_info['file_path']),
                                'dimension': dim_name,
                                'line': line_num,
                                'field': field_name,
                                'type': field_type,
                                'comment': f'Field "{field_name}" of type "{field_type}" is missing @AMetaData annotation'
                            })
                    elif dim_name == "Validation Annotations":
                        missing = [v for v in field_violations if v[0] == 'missing_validation']
                        for violation_type, field_name, line_num, field_type in missing:
                            detailed_violations.append({
                                'class': class_info['class_name'],
                                'file': str(class_info['file_path']),
                                'dimension': dim_name,
                                'line': line_num,
                                'field': field_name,
                                'type': field_type,
                                'comment': f'Field "{field_name}" of type "{field_type}" is missing validation annotations (@NotNull, @NotBlank, or @Size)'
                            })
        
        row += 1
    
    # Add Violations Detail sheet (for classes with 10+ violations)
    if detailed_violations:
        violations_sheet = wb.create_sheet("Violations Detail")
        
        # Header
        violations_sheet.cell(1, 1, "Detailed Violations Report").font = Font(size=16, bold=True)
        violations_sheet.cell(2, 1, f"Classes with 10+ violations: {len(set(v['class'] for v in detailed_violations))} classes").font = Font(size=12)
        violations_sheet.cell(3, 1, f"Total violations listed: {len(detailed_violations)}").font = Font(size=12)
        
        # Column headers
        headers_row = 5
        violation_headers = ["Class Name", "File Path", "Quality Dimension", "Line Number", "Field Name", "Field Type", "Comment/Description"]
        for col_idx, header in enumerate(violation_headers, 1):
            cell = violations_sheet.cell(headers_row, col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Set column widths
        violations_sheet.column_dimensions['A'].width = 30  # Class Name
        violations_sheet.column_dimensions['B'].width = 50  # File Path
        violations_sheet.column_dimensions['C'].width = 25  # Quality Dimension
        violations_sheet.column_dimensions['D'].width = 12  # Line Number
        violations_sheet.column_dimensions['E'].width = 25  # Field Name
        violations_sheet.column_dimensions['F'].width = 20  # Field Type
        violations_sheet.column_dimensions['G'].width = 60  # Comment
        
        # Freeze panes (freeze header rows)
        violations_sheet.freeze_panes = 'A6'
        
        # Add violation details
        detail_row = headers_row + 1
        for violation in sorted(detailed_violations, key=lambda x: (x['class'], x['line'])):
            violations_sheet.cell(detail_row, 1, violation['class']).border = border
            violations_sheet.cell(detail_row, 2, violation['file']).border = border
            violations_sheet.cell(detail_row, 3, violation['dimension']).border = border
            violations_sheet.cell(detail_row, 4, violation['line']).border = border
            violations_sheet.cell(detail_row, 5, violation['field']).border = border
            violations_sheet.cell(detail_row, 6, violation['type']).border = border
            cell = violations_sheet.cell(detail_row, 7, violation['comment'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
            detail_row += 1
        
        print(f"Added {len(detailed_violations)} detailed violations to 'Violations Detail' sheet")
    
    # Add summary sheet
    summary = wb.create_sheet("Summary")
    summary.cell(1, 1, "Code Quality Matrix Summary").font = Font(size=16, bold=True)
    summary.cell(3, 1, "Total Classes:").font = Font(bold=True)
    summary.cell(3, 2, len(class_paths))
    summary.cell(4, 1, "Quality Dimensions:").font = Font(bold=True)
    summary.cell(4, 2, len(QUALITY_DIMENSIONS))
    
    summary.cell(6, 1, "Legend:").font = Font(bold=True)
    summary.cell(7, 1, "0").fill = zero_fill
    summary.cell(7, 1).font = Font(color="006100", bold=True)
    summary.cell(7, 2, "No violations - 100% compliance (perfect)")
    
    summary.cell(8, 1, "1-3").fill = low_fill
    summary.cell(8, 1).font = Font(color="9C5700")
    summary.cell(8, 2, "Low violations - Minor issues to address")
    
    summary.cell(9, 1, "4-6").fill = medium_fill
    summary.cell(9, 1).font = Font(color="9C2700")
    summary.cell(9, 2, "Medium violations - Multiple issues need attention")
    
    summary.cell(10, 1, "7+").fill = high_fill
    summary.cell(10, 1).font = Font(color="9C0006", bold=True)
    summary.cell(10, 2, "High violations - Significant quality issues")
    
    summary.cell(11, 1, "-").fill = na_fill
    summary.cell(11, 2, "N/A - Quality gate not applicable to this class")
    
    summary.cell(13, 1, "Notes:").font = Font(bold=True)
    summary.cell(14, 1, "• Violation counts are capped at 10 (shown as '10+')")
    summary.cell(15, 1, "• 0 violations = 100% compliance with quality gate")
    summary.cell(16, 1, "• Use TODO comments in source files to mark violations")
    summary.cell(17, 1, "• TODO format: // TODO: [Gate Name] - [Issue] - [Fix]")
    
    summary.column_dimensions['A'].width = 15
    summary.column_dimensions['B'].width = 60
    
    # Save workbook
    wb.save(output_file)
    print(f"\nExcel matrix saved to: {output_file}")
    print(f"Total rows: {row - 1}")
    print(f"Total quality dimensions: {len(QUALITY_DIMENSIONS)}")

if __name__ == "__main__":
    classes_file = "/tmp/quality_matrix/all_classes.txt"
    output_file = BASE_DIR / "docs" / "CODE_QUALITY_MATRIX.xlsx"
    
    create_excel_matrix(classes_file, str(output_file))
